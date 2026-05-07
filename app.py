# app_landuse.py
# PlanVision AI v2 — 토지이용계획도 기반 도시개발 조감도 자동생성
# 입력: 토지이용계획도 + 위성사진(선택) + 토지이용계획표 (RGB 매핑)
# 3-STEP: PASS1(2D배치도) → PASS2(3D조감도) → PASS3(각도변환)

import tempfile
import os
import json
import math
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from io import BytesIO

import streamlit as st
from PIL import Image, ImageDraw, ImageFilter

try:
    from streamlit_folium import st_folium
except Exception:
    st_folium = None

try:
    from pyproj import Transformer
    PYPROJ_AVAILABLE = True
except Exception:
    PYPROJ_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except Exception:
    REQUESTS_AVAILABLE = False

try:
    import folium
except Exception:
    folium = None

# ── 선택적 의존성 ──────────────────────────────────────────────
GENAI_AVAILABLE = True
GENAI_TYPES_AVAILABLE = True
try:
    from google import genai
    try:
        from google.genai import types  # type: ignore
    except Exception:
        GENAI_TYPES_AVAILABLE = False
except Exception:
    GENAI_AVAILABLE = False
    GENAI_TYPES_AVAILABLE = False

CV2_AVAILABLE = True
try:
    import cv2  # type: ignore
    import numpy as np  # type: ignore
except Exception:
    CV2_AVAILABLE = False
    np = None  # type: ignore

EZDXF_AVAILABLE = True
try:
    import ezdxf
    from ezdxf.colors import aci2rgb
except Exception:
    EZDXF_AVAILABLE = False

# ──────────────────────────────────────────────────────────────
# 페이지 설정
# ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="PlanVision AI v2 — 토지이용계획 기반", layout="wide")

st.markdown("""
<style>
:root {
    --primary: #2563EB;
    --primary-light: #EFF6FF;
    --success: #16A34A;
    --border: #E5E7EB;
    --radius: 10px;
}
.block-container { max-width: 1280px; padding: 1.5rem 2rem; margin: 0 auto; }
div[data-testid="stButton"] button { border-radius:8px !important; font-weight:600 !important; }
div[data-testid="stButton"] button[kind="primary"] {
    background:#2563EB !important; border:none !important; color:#fff !important;
}
.section-header {
    font-size: 20px; font-weight: 800; color: #111827;
    border-left: 4px solid #2563EB;
    padding: 10px 16px; background: #EFF6FF;
    border-radius: 0 8px 8px 0; margin: 20px 0 12px 0;
}
.sub-label {
    font-size: 11px; font-weight: 700; color: #9CA3AF;
    letter-spacing: 0.1em; text-transform: uppercase;
    margin: 16px 0 6px 0;
}
[data-testid="stDecoration"] { display:none !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="padding:16px 0 20px 0;">
  <div style="font-size:13px;font-weight:700;color:#2563EB;letter-spacing:.1em;text-transform:uppercase;margin-bottom:6px;">
    KH ENGINEERING · URBAN AI
  </div>
  <div style="font-size:40px;font-weight:900;letter-spacing:-.02em;
    background:linear-gradient(90deg,#111827,#1E40AF);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;line-height:1.1;">
    PlanVision AI <span style="font-size:18px;font-weight:600;opacity:.5;">v2</span>
  </div>
  <div style="font-size:14px;color:#6B7280;margin-top:6px;">
    토지이용계획도 + RGB 매핑 기반 도시개발 조감도 자동생성
  </div>
</div>
<div style="height:3px;background:linear-gradient(90deg,#2563EB,#1E293B 80%);border-radius:2px;margin-bottom:28px;"></div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────────────────────
MODEL_NAME = "gemini-3.1-flash-image-preview"

STANDARD_LAND_USES = [
    # 주거
    ("단독주택",          (255, 255, 127), "단독주택",                    "저층 단독주택"),
    ("연립·다세대주택",   (255, 230, 100), "연립·다세대주택",             "저~중층 연립"),
    ("공동주택(판상)",    (255, 220,  60), "공동주택(판상형 아파트)",     "판상형 아파트"),
    ("공동주택(타워)",    (255, 180,  30), "공동주택(타워형 아파트)",     "타워형 아파트"),
    ("준주거용지",        (255, 200, 150), "준주거용지",                  "준주거 혼합"),
    ("숙박시설",          (255, 191, 127), "[직접입력]",                  ""),
    ("콘도미니엄",        (159, 127, 255), "[직접입력]",                  ""),
    # 상업·업무
    ("근린생활시설",      (255, 140, 140), "근린생활시설용지",            "근린상가"),
    ("일반상업용지",      (255,  80,  80), "일반상업용지",                "일반상업"),
    ("복합상업시설",      (220,  50,  50), "복합상업시설(대형몰·복합몰)", "복합상업"),
    ("업무시설용지",      (180, 130, 220), "업무시설용지(오피스)",        "오피스"),
    ("복합업무용지",      (150, 100, 200), "복합업무용지(오피스+상업)",   "복합업무"),
    ("R&D·첨단산업",     (130, 100, 180), "첨단산업단지(R&D·지식산업)",  "R&D 캠퍼스"),
    ("6차산업",           (165,  82, 124), "[직접입력]",                  ""),
    ("스마트팜",          (223, 127, 255), "[직접입력]",                  ""),
    ("파머스마켓",        (255,   0, 255), "[직접입력]",                  ""),
    # 공원·녹지
    ("공원",              (  0, 165,   0), "근린공원·주제공원",           "공원녹지"),
    ("녹지·완충녹지",     (191, 255, 127), "근린공원·주제공원",           "완충녹지"),
    ("치유의숲",          (127, 255,   0), "[직접입력]",                  ""),
    ("마을농원",          (145, 165,  82), "[직접입력]",                  ""),
    ("파크골프장",        (103, 165,  82), "[직접입력]",                  ""),
    ("하천·수변",         ( 80, 160, 220), "하천·수변공간",               "수변공간"),
    ("저류지",            (127, 223, 255), "[직접입력]",                  ""),
    ("인피니티풀",        (173, 241, 255), "[직접입력]",                  ""),
    # 공공·편의
    ("공공청사",          (180, 220, 200), "공공청사·행정시설",           "공공청사"),
    ("학교·교육",         (200, 230, 255), "학교·교육시설",               "교육시설"),
    ("의료시설",          (220, 240, 255), "종합의료시설(병원)",          "병원"),
    ("문화시설",          (220, 180, 240), "대규모 문화시설(공연·전시·컨벤션)", "문화시설"),
    ("복지시설",          (165, 165,   0), "[직접입력]",                  ""),
    ("주민편의시설",      (127, 191, 255), "[직접입력]",                  ""),
    ("복합커뮤니티시설",  (255, 159, 127), "[직접입력]",                  ""),
    ("복합문화체육시설",  ( 82, 165, 124), "[직접입력]",                  ""),
    ("버스킹공연장",      (251, 203, 229), "[직접입력]",                  ""),
    # 기반시설
    ("광장·공공공지",     (240, 240, 240), "광장·공공공지",               "광장"),
    ("주차장",            (137, 137, 137), "주차장",                      "주차장"),
    ("도로",              (255, 255, 255), "광장·공공공지",               "도로"),
    ("보행자전용도로",    (165, 124,   0), "[직접입력]",                  ""),
    ("산책로",            (165,  82,   0), "[직접입력]",                  ""),
    # 산업·물류
    ("일반산업단지",      (200, 170, 130), "일반산업단지(공장용지)",      "공장용지"),
    ("물류단지",          (180, 150, 110), "첨단물류단지",                "물류"),
    ("복합용지",          (255, 200, 180), "복합용지(혼합개발)",          "혼합개발"),
]

# ──────────────────────────────────────────────────────────────
# 프리셋
# ──────────────────────────────────────────────────────────────
ZONE_PRESETS_SIMPLE = {
    "단독주택": {
        "Primary Function": "Residential - Detached housing",
        "floor_level": "Low",
        "prompt_note": "Low-rise detached housing, 2~3F, garden plots, private residential character",
    },
    "연립·다세대주택": {
        "Primary Function": "Residential - Attached housing",
        "floor_level": "Low",
        "prompt_note": "Low-to-mid rise attached housing, 3~5F, courtyard arrangement",
    },
    "공동주택(판상형 아파트)": {
        "Primary Function": "Residential - Slab apartment complex",
        "floor_level": "Medium–High",
        "prompt_note": "Slab-type apartment, 8~15F, south-facing orientation, central green space",
    },
    "공동주택(타워형 아파트)": {
        "Primary Function": "Residential - High-rise tower apartment",
        "floor_level": "High",
        "prompt_note": "High-rise point tower apartment, 20~30F, large landscaped podium",
    },
    "준주거용지": {
        "Primary Function": "Quasi-residential mixed-use",
        "floor_level": "Low",
        "prompt_note": "Mixed-use quasi-residential, ground retail with residential above, 3~6F",
    },
    "근린생활시설용지": {
        "Primary Function": "Commercial dominant",
        "floor_level": "Low",
        "prompt_note": "Low-rise neighborhood commercial strip, 2~4F storefronts",
    },
    "일반상업용지": {
        "Primary Function": "Commercial dominant",
        "floor_level": "Medium–High",
        "prompt_note": "General commercial zone, podium-and-tower typology, 8~20F",
    },
    "복합상업시설(대형몰·복합몰)": {
        "Primary Function": "Commercial dominant",
        "floor_level": "Medium–High",
        "prompt_note": "Large-scale mixed commercial complex, COEX style, 10~25F",
    },
    "업무시설용지(오피스)": {
        "Primary Function": "Office dominant",
        "floor_level": "High",
        "prompt_note": "Office tower district, plaza-level retail activation, 15~30F",
    },
    "복합업무용지(오피스+상업)": {
        "Primary Function": "Office + Commercial mixed-use",
        "floor_level": "High",
        "prompt_note": "Mixed office and commercial complex, podium retail with tower, 15~30F",
    },
    "첨단산업단지(R&D·지식산업)": {
        "Primary Function": "Innovation/R&D dominant",
        "floor_level": "Medium",
        "prompt_note": "High-tech R&D campus, courtyard green network, 3~8F",
    },
    "근린공원·주제공원": {
        "Primary Function": "Park / Open space",
        "floor_level": "Very low",
        "prompt_note": "Neighborhood park, tree canopy, walking paths, event lawn, no buildings",
    },
    "하천·수변공간": {
        "Primary Function": "Park / Open space",
        "floor_level": "Very low",
        "prompt_note": "River corridor, riparian planting, boardwalk, water feature visible, no buildings",
    },
    "공공청사·행정시설": {
        "Primary Function": "Civic / Government",
        "floor_level": "Medium",
        "prompt_note": "Civic government building, formal plaza entry, institutional character, 5~12F",
    },
    "학교·교육시설": {
        "Primary Function": "Education",
        "floor_level": "Low",
        "prompt_note": "School campus, playgrounds and sports fields, 2~4F",
    },
    "종합의료시설(병원)": {
        "Primary Function": "Medical / Healthcare",
        "floor_level": "Medium–High",
        "prompt_note": "General hospital complex, tower block with podium, 8~20F",
    },
    "대규모 문화시설(공연·전시·컨벤션)": {
        "Primary Function": "Large-scale cultural / Convention",
        "floor_level": "Medium",
        "prompt_note": "Large cultural landmark: convention center, grand civic plaza",
    },
    "광장·공공공지": {
        "Primary Function": "Open space dominant",
        "floor_level": "Very low",
        "prompt_note": "Civic plaza, paved surface, fountain or public art, no buildings",
    },
    "주차장": {
        "Primary Function": "Parking",
        "floor_level": "Very low",
        "prompt_note": "Surface or structured parking lot, organized layout, 1~3F",
    },
    "일반산업단지(공장용지)": {
        "Primary Function": "General industrial",
        "floor_level": "Low",
        "prompt_note": "General industrial zone, large-footprint factory buildings, low-rise",
    },
    "첨단물류단지": {
        "Primary Function": "Logistics / Distribution",
        "floor_level": "Low",
        "prompt_note": "Advanced logistics center, large-scale warehouse buildings, truck access roads",
    },
    "복합용지(혼합개발)": {
        "Primary Function": "Mixed urban fabric",
        "floor_level": "Medium–High",
        "prompt_note": "Mixed-use development zone, podium base with tower elements",
    },
}

PARK_PF = "Park / Open space"
PRESET_KEYS = list(ZONE_PRESETS_SIMPLE.keys())
PRESET_OPTIONS = ["[직접입력]"] + PRESET_KEYS

# ──────────────────────────────────────────────────────────────
# 세션 초기화
# ──────────────────────────────────────────────────────────────
def ensure_session():
    defs = {
        "step": 0,
        "img_landuse_bytes": None,
        "img_sat_bytes": None,
        "land_use_table": [],
        "site_area_sqm": 100000.0,
        "pass1_outputs": [],
        "pass1_selected_idx": 0,
        "pass1_output_bytes": None,
        "pass2_outputs": [],
        "pass2_selected_idx": 0,
        "pass2_output_bytes": None,
        "pass3_outputs": [],
        "pass3_selected_idx": 0,
        "_auto_colors": [],
        "_auto_generated": False,
        "_errors": [],
        "dxf_bytes": None,
        "dxf_layer_table": [],
        "use_dxf_mapping": True,
        "dxf_preview_bytes": None,
        "dxf_satellite_base_bytes": None,
        "dxf_landuse_hatch_bytes": None,
        "dxf_bbox": None,
        "dxf_crs": "EPSG:5174",
    }
    for k, v in defs.items():
        if k not in st.session_state:
            st.session_state[k] = v

ensure_session()

# DXF/VWorld 추출 단계에서 쓰는 추가 세션값
for _k, _v in {
    "dxf_rows": [],
    "dxf_geojson": None,
    "gdal_geojson": None,
    "hatch_area_map": {},
    "records": [],
    "preview_png": None,
    "satellite_base_png": None,
    "landuse_hatch_png": None,
    "bbox_3857": None,
    "prompt_text": "",
    "export_bbox_3857": None,
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ──────────────────────────────────────────────────────────────
# 기본 테이블
# ──────────────────────────────────────────────────────────────
def make_default_table():
    defaults = [
        ("단독주택",         (255, 255, 127), "단독주택",               "",                                                                  10000.0),
        ("공원",             (  0, 165,   0), "근린공원·주제공원",      "",                                                                  15000.0),
        ("녹지·완충녹지",    (191, 255, 127), "근린공원·주제공원",      "",                                                                   8000.0),
        ("공공청사",         (180, 220, 200), "공공청사·행정시설",      "",                                                                   5000.0),
        ("주차장",           (137, 137, 137), "주차장",                 "",                                                                   3000.0),
        ("숙박시설",         (255, 191, 127), "[직접입력]",             "resort hotel with amenity facilities, 5~10F, warm facade",          12000.0),
        ("복합커뮤니티시설", (255, 159, 127), "[직접입력]",             "community center with multipurpose hall and outdoor plaza",           4000.0),
        ("치유의숲",         (127, 255,   0), "[직접입력]",             "healing forest with walking trails, meditation zones, no buildings", 20000.0),
    ]
    return [
        {
            "name": n,
            "r": rgb[0], "g": rgb[1], "b": rgb[2],
            "preset": preset,
            "custom_desc": custom_desc,
            "area_sqm": area,
            "tolerance": 25,
            "enabled": True,
        }
        for n, rgb, preset, custom_desc, area in defaults
    ]

if not st.session_state.land_use_table:
    st.session_state.land_use_table = []

# ──────────────────────────────────────────────────────────────
# 이미지 유틸
# ──────────────────────────────────────────────────────────────
def pil_to_png_bytes(img: Image.Image) -> bytes:
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def bytes_to_pil(b: bytes) -> Image.Image:
    return Image.open(BytesIO(b)).convert("RGB")


# ──────────────────────────────────────────────────────────────
# DXF/VWorld 고정밀 PNG 추출 유틸 — app_dxf_vworld_mvp.py 통합
# ──────────────────────────────────────────────────────────────
# -----------------------------
# Constants
# -----------------------------
CRS_OPTIONS = {
    "EPSG:5174 - Korean 1985 / Modified Central Belt": 5174,
    "EPSG:5179 - Korea 2000 / Unified CS": 5179,
    "EPSG:5186 - Korea 2000 / Central Belt": 5186,
    "EPSG:5187 - Korea 2000 / East Belt": 5187,
    "EPSG:4326 - WGS84": 4326,
}

IMG_SIZE_OPTIONS = {
    "1024 x 768": (1024, 768),
    "1280 x 960": (1280, 960),
    "1600 x 1200": (1600, 1200),
    "2048 x 1536": (2048, 1536),
}

# VWorld WMTS/TMS tile URL pattern.
# Satellite uses jpeg; Base/Hybrid use png in most examples.
VWORLD_API_KEY = "05CD1D67-6203-3E82-BFDA-BBC5DE6AA857"
VWORLD_SAT_URL = "https://api.vworld.kr/req/wmts/1.0.0/{key}/Satellite/{z}/{y}/{x}.jpeg"
VWORLD_BASE_URL = "https://api.vworld.kr/req/wmts/1.0.0/{key}/Base/{z}/{y}/{x}.png"

DEFAULT_MAP_ZOOM = 17
DEFAULT_EXPORT_ZOOM = 19
DEFAULT_EXPORT_SCALE = 2

# -----------------------------
# Layer rules
# -----------------------------
def clean_layer_name(layer_name: str) -> str:
    return str(layer_name).replace(" ", "").replace("_", "").replace("-", "").replace(".", "").lower()


def is_boundary_layer(layer_name: str) -> bool:
    n = clean_layer_name(layer_name)
    return "구역계" in n or "boundary" in n or "사업대상지" in n


def is_plan_line_layer(layer_name: str) -> bool:
    n = clean_layer_name(layer_name)
    plan_keywords = [
        "계획선",
        "획지선",
        "가구선",
        "도로선",
        "중심선",
        "planline",
        "parcelline",
    ]
    return any(k in n for k in plan_keywords)


def is_landuse_layer(layer_name: str) -> bool:
    n = clean_layer_name(layer_name)
    if is_boundary_layer(layer_name) or is_plan_line_layer(layer_name):
        return False
    keys = [
        "h공원", "h녹지", "h도로", "h도시지원시설", "h보행자전용도로",
        "h복합용지", "h산업", "h산업복합", "h산업지원시설", "h소하천", "h주차장",
        "공원", "녹지", "도로", "도시지원시설", "복합", "산업", "상업", "주차", "하천", "수변", "주거"
    ]
    return any(k in n for k in keys)


def guess_program_from_layer(layer_name: str) -> str:
    n = clean_layer_name(layer_name)
    if "공원" in n:
        return "landscape park / neighborhood park, no buildings"
    if "녹지" in n:
        return "green buffer / landscape open space, no buildings"
    if "하천" in n or "소하천" in n or "수변" in n:
        return "river / waterfront corridor, riparian landscape, no buildings"
    if "도로" in n:
        return "road network / asphalt roads / pedestrian and vehicle circulation"
    if "주차" in n:
        return "parking facility / surface or structured parking"
    if "상업" in n:
        return "commercial / mixed-use commercial district"
    if "복합" in n:
        return "mixed-use development zone, podium and tower composition"
    if "도시지원" in n or "산업지원" in n:
        return "urban support facility / business support / civic-industrial support"
    if "산업" in n:
        return "business, R&D and light industrial campus"
    if "주거" in n:
        return "residential blocks with landscaped courtyards"
    return "user-defined land use program"

# -----------------------------
# DXF utilities
# -----------------------------
def aci_to_rgb(aci: int):
    try:
        rgb = aci2rgb(int(aci))
        return (int(rgb.r), int(rgb.g), int(rgb.b))
    except Exception:
        return (180, 180, 180)


def true_color_to_rgb(true_color: int):
    r = (true_color >> 16) & 255
    g = (true_color >> 8) & 255
    b = true_color & 255
    return (r, g, b)


def entity_rgb(doc, e):
    try:
        if e.dxf.hasattr("true_color") and e.dxf.true_color is not None:
            return true_color_to_rgb(e.dxf.true_color)
        if e.dxf.hasattr("color") and int(e.dxf.color) not in [0, 256]:
            return aci_to_rgb(int(e.dxf.color))
        layer = doc.layers.get(e.dxf.layer)
        if layer.dxf.hasattr("true_color") and layer.dxf.true_color is not None:
            return true_color_to_rgb(layer.dxf.true_color)
        return aci_to_rgb(int(layer.dxf.color))
    except Exception:
        return (180, 180, 180)


def write_temp_dxf(dxf_bytes: bytes) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".dxf")
    tmp.write(dxf_bytes)
    tmp.close()
    return tmp.name


def find_ogr2ogr():
    exe = shutil.which("ogr2ogr")
    if exe:
        return exe
    candidates = [
        r"C:\Program Files\QGIS 3.44.1\bin\ogr2ogr.exe",
        r"C:\Program Files\QGIS 3.44.0\bin\ogr2ogr.exe",
        r"C:\Program Files\QGIS 3.40.0\bin\ogr2ogr.exe",
        r"C:\Program Files\QGIS 3.38.0\bin\ogr2ogr.exe",
        r"C:\Program Files\QGIS 3.36.0\bin\ogr2ogr.exe",
        r"C:\Program Files\QGIS 3.34.0\bin\ogr2ogr.exe",
        r"C:\OSGeo4W\bin\ogr2ogr.exe",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p

    # QGIS 버전 폴더 자동 스캔
    qgis_root = r"C:\Program Files"
    try:
        for name in os.listdir(qgis_root):
            if name.startswith("QGIS "):
                p = os.path.join(qgis_root, name, "bin", "ogr2ogr.exe")
                if os.path.exists(p):
                    return p
    except Exception:
        pass

    raise RuntimeError(
        "ogr2ogr.exe를 자동으로 찾지 못했습니다. "
        "QGIS가 C:\\Program Files\\QGIS x.xx.x 경로에 설치되어 있는지 확인하세요."
    )


def dxf_to_geojson_by_gdal(dxf_bytes: bytes, src_epsg: int):
    """GDAL/OGR 방식으로 표시용 WGS84 GeoJSON을 생성."""
    ogr2ogr = find_ogr2ogr()
    with tempfile.TemporaryDirectory() as td:
        dxf_path = os.path.join(td, "input.dxf")
        display_path = os.path.join(td, "display_4326.geojson")
        with open(dxf_path, "wb") as f:
            f.write(dxf_bytes)

        cmd_display = [
            ogr2ogr,
            "-f", "GeoJSON",
            display_path,
            dxf_path,
            "-s_srs", f"EPSG:{src_epsg}",
            "-t_srs", "EPSG:4326",
            "-lco", "RFC7946=YES",
            "-nlt", "PROMOTE_TO_MULTI",
            "-skipfailures",
        ]

        result = subprocess.run(cmd_display, capture_output=True, text=True, encoding="utf-8", errors="ignore")
        if result.returncode != 0:
            raise RuntimeError(result.stderr or result.stdout or "ogr2ogr 변환 실패")

        with open(display_path, "r", encoding="utf-8") as f:
            return json.load(f)


def build_layer_rgb_map_from_dxf(dxf_bytes: bytes):
    """DXF 원본에서 레이어별 대표 RGB 추출. GDAL geometry + ezdxf color 조합용."""
    if not EZDXF_AVAILABLE:
        return {}
    tmp_path = write_temp_dxf(dxf_bytes)
    try:
        doc = ezdxf.readfile(tmp_path)
        msp = doc.modelspace()
        counter = defaultdict(Counter)
        for e in msp:
            try:
                layer = e.dxf.layer
                if not (is_landuse_layer(layer) or is_boundary_layer(layer) or is_plan_line_layer(layer)):
                    continue
                rgb = entity_rgb(doc, e)
                counter[layer][rgb] += 1
            except Exception:
                continue
        return {layer: c.most_common(1)[0][0] for layer, c in counter.items() if c}
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def iter_geojson_geometries(geom):
    """GeoJSON geometry를 (geom_type, coords) 단위로 분해."""
    if not geom:
        return
    gtype = geom.get("type")
    coords = geom.get("coordinates", [])
    if gtype == "Polygon":
        yield "polygon", coords
    elif gtype == "MultiPolygon":
        for poly in coords:
            yield "polygon", poly
    elif gtype == "LineString":
        yield "line", coords
    elif gtype == "MultiLineString":
        for line in coords:
            yield "line", line
    elif gtype == "GeometryCollection":
        for g in geom.get("geometries", []):
            yield from iter_geojson_geometries(g)


def get_feature_layer_name(props):
    """GDAL DXF driver가 넣는 레이어명 필드 후보 처리."""
    for key in ["Layer", "layer", "LAYER", "Name", "name"]:
        if key in props and props[key]:
            return str(props[key])
    return ""


def records_from_gdal_geojson(geojson, dxf_bytes: bytes, src_epsg: int):
    """GDAL GeoJSON(EPSG:4326)을 draw_records용 EPSG:3857 records로 변환."""
    t = get_transformer(4326, 3857)
    layer_rgb_map = build_layer_rgb_map_from_dxf(dxf_bytes)
    records = []
    xs, ys = [], []
    for feat in geojson.get("features", []):
        props = feat.get("properties", {}) or {}
        layer = get_feature_layer_name(props)
        if not layer:
            continue
        category = "other"
        if is_boundary_layer(layer):
            category = "boundary"
        elif is_landuse_layer(layer):
            category = "landuse"
        elif is_plan_line_layer(layer):
            category = "line"
        if category == "other":
            continue
        rgb = layer_rgb_map.get(layer, (180, 180, 180))
        for geom_type, coords in iter_geojson_geometries(feat.get("geometry")):
            if geom_type == "polygon":
                if not coords:
                    continue
                outer_ll = coords[0]
                holes_ll = coords[1:] if len(coords) > 1 else []
                outer_3857 = []
                for lon, lat, *_rest in outer_ll:
                    x, y = transform_xy(lon, lat, t)
                    outer_3857.append((x, y))
                    xs.append(x); ys.append(y)
                holes_3857 = []
                for hole in holes_ll:
                    h = []
                    for lon, lat, *_rest in hole:
                        x, y = transform_xy(lon, lat, t)
                        h.append((x, y))
                    if len(h) >= 4:
                        holes_3857.append(h)
                if len(outer_3857) >= 4:
                    records.append({
                        "layer": layer, "type": "GDAL_POLYGON", "category": category,
                        "rgb": rgb, "points": outer_3857,
                        "holes": holes_3857, "geom_type": "polygon",
                    })
            elif geom_type == "line":
                pts_3857 = []
                for lon, lat, *_rest in coords:
                    x, y = transform_xy(lon, lat, t)
                    pts_3857.append((x, y))
                    xs.append(x); ys.append(y)
                if len(pts_3857) >= 2:
                    records.append({
                        "layer": layer, "type": "GDAL_LINE", "category": category,
                        "rgb": rgb, "points": pts_3857,
                        "holes": [], "geom_type": "line",
                    })
    bbox = (min(xs), min(ys), max(xs), max(ys)) if xs and ys else None
    return records, bbox


def get_transformer(src_epsg: int, dst_epsg: int):
    if src_epsg == dst_epsg:
        return None
    return Transformer.from_crs(f"EPSG:{src_epsg}", f"EPSG:{dst_epsg}", always_xy=True)


def transform_xy(x, y, transformer):
    if transformer is None:
        return float(x), float(y)
    return transformer.transform(float(x), float(y))


def close_ring(points):
    pts = list(points)
    if pts and pts[0] != pts[-1]:
        pts.append(pts[0])
    return pts


def polygon_area(points):
    if len(points) < 3:
        return 0.0
    area = 0.0
    pts = points[:]
    if pts[0] != pts[-1]:
        pts.append(pts[0])
    for (x1, y1), (x2, y2) in zip(pts[:-1], pts[1:]):
        area += x1 * y2 - x2 * y1
    return abs(area) * 0.5


def sample_arc_edge_safe(edge, flatten_distance=0.01):
    cx, cy = edge.center
    radius = float(edge.radius)
    a0 = math.radians(float(edge.start_angle))
    a1 = math.radians(float(edge.end_angle))
    if a1 < a0:
        a1 += math.tau
    arc_len = abs(a1 - a0) * radius
    segs = max(16, min(4096, int(arc_len / max(flatten_distance, 0.001))))
    pts = []
    for i in range(segs + 1):
        a = a0 + (a1 - a0) * i / segs
        pts.append((float(cx + radius * math.cos(a)), float(cy + radius * math.sin(a))))
    if hasattr(edge, "ccw") and not edge.ccw:
        pts.reverse()
    return pts


def compute_hatch_area_by_layer(dxf_bytes):
    import ezdxf
    from collections import defaultdict

    tmp_path = write_temp_dxf(dxf_bytes)
    area_map = defaultdict(float)

    try:
        doc = ezdxf.readfile(tmp_path)
        msp = doc.modelspace()

        for e in msp:
            if e.dxftype() != "HATCH":
                continue

            layer = e.dxf.layer

            if not is_landuse_layer(layer):
                continue

            total_area = 0.0

            for path in e.paths:
                if hasattr(path, "vertices"):
                    pts = []
                    for v in path.vertices:
                        x = float(v[0])
                        y = float(v[1])
                        pts.append((x, y))

                    if len(pts) >= 3:
                        total_area += polygon_area(close_ring(pts))

                elif hasattr(path, "edges"):
                    pts = []
                    for edge in path.edges:
                        if edge.EDGE_TYPE == "LineEdge":
                            if not pts:
                                pts.append((float(edge.start[0]), float(edge.start[1])))
                            pts.append((float(edge.end[0]), float(edge.end[1])))

                        elif edge.EDGE_TYPE == "ArcEdge":
                            arc_pts = sample_arc_edge_safe(edge, flatten_distance=0.01)
                            if pts and arc_pts:
                                pts.extend(arc_pts[1:])
                            else:
                                pts.extend(arc_pts)

                    if len(pts) >= 3:
                        total_area += polygon_area(close_ring(pts))

            area_map[layer] += abs(total_area)

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    return dict(area_map)


def build_dxf_layer_table(records, area_map=None):
    area_map = area_map or {}
    grouped = {}

    for rec in records:
        if rec["category"] != "landuse":
            continue

        layer = rec["layer"]

        if layer not in grouped:
            grouped[layer] = {
                "enabled": True,
                "layer": layer,
                "color_counter": Counter(),
                "entity_count": 0,
            }

        grouped[layer]["entity_count"] += 1
        grouped[layer]["color_counter"][rec["rgb"]] += 1

    rows = []
    for layer, g in grouped.items():
        rgb = g["color_counter"].most_common(1)[0][0] if g["color_counter"] else (180, 180, 180)
        r, gc, b = rgb

        rows.append({
            "enabled": True,
            "layer": layer,
            "rgb": f"RGB({r},{gc},{b})",
            "area_sqm": round(float(area_map.get(layer, 0.0)), 1),
        })

    return sorted(rows, key=lambda r: r["layer"])


# -----------------------------
# Tile utilities
# -----------------------------
def mercator_to_tile(x, y, z):
    origin_shift = 20037508.342789244
    n = 2 ** z
    tx = int((x + origin_shift) / (2 * origin_shift) * n)
    ty = int((origin_shift - y) / (2 * origin_shift) * n)
    return tx, ty


def tile_bounds_mercator(x, y, z):
    origin_shift = 20037508.342789244
    n = 2 ** z
    tile_size = 2 * origin_shift / n
    minx = x * tile_size - origin_shift
    maxx = (x + 1) * tile_size - origin_shift
    maxy = origin_shift - y * tile_size
    miny = origin_shift - (y + 1) * tile_size
    return minx, miny, maxx, maxy


def fetch_vworld_tile(key, z, x, y, layer="Satellite"):
    if not REQUESTS_AVAILABLE or not key:
        return None
    if layer == "Satellite":
        url = VWORLD_SAT_URL.format(key=key, z=z, x=x, y=y)
    else:
        url = VWORLD_BASE_URL.format(key=key, z=z, x=x, y=y)
    try:
        resp = requests.get(url, timeout=8, headers={"User-Agent":"PlanVision/1.0"})
        if resp.status_code == 200 and resp.content:
            return Image.open(BytesIO(resp.content)).convert("RGB")
    except Exception:
        return None
    return None


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def fetch_vworld_tile_cached(key, z, x, y, layer="Satellite"):
    return fetch_vworld_tile(key, z, x, y, layer)


def expand_bbox_to_aspect(bbox3857, target_w, target_h, padding_ratio=0.05):
    minx, miny, maxx, maxy = bbox3857

    dx = maxx - minx
    dy = maxy - miny

    pad = max(dx, dy) * padding_ratio
    minx -= pad
    maxx += pad
    miny -= pad
    maxy += pad

    dx = maxx - minx
    dy = maxy - miny

    target_aspect = target_w / target_h
    bbox_aspect = dx / dy

    cx = (minx + maxx) / 2
    cy = (miny + maxy) / 2

    if bbox_aspect > target_aspect:
        # bbox가 너무 넓음 -> 세로 확장
        new_dy = dx / target_aspect
        miny = cy - new_dy / 2
        maxy = cy + new_dy / 2
    else:
        # bbox가 너무 높음 -> 가로 확장
        new_dx = dy * target_aspect
        minx = cx - new_dx / 2
        maxx = cx + new_dx / 2

    return minx, miny, maxx, maxy


def make_satellite_mosaic(bbox3857, key, z=18, width=1600, height=1200, padding_ratio=0.05):
    minx, miny, maxx, maxy = expand_bbox_to_aspect(
        bbox3857,
        width,
        height,
        padding_ratio=padding_ratio,
    )

    tx_min, ty_max = mercator_to_tile(minx, miny, z)
    tx_max, ty_min = mercator_to_tile(maxx, maxy, z)

    tile_size = 256
    mosaic_w = (tx_max - tx_min + 1) * tile_size
    mosaic_h = (ty_max - ty_min + 1) * tile_size

    mosaic = Image.new("RGB", (mosaic_w, mosaic_h), (245, 245, 245))

    for tx in range(tx_min, tx_max + 1):
        for ty in range(ty_min, ty_max + 1):
            tile = fetch_vworld_tile_cached(key, z, tx, ty, layer="Satellite")
            if tile is None:
                tile = Image.new("RGB", (tile_size, tile_size), (235, 235, 235))
            mosaic.paste(tile.resize((tile_size, tile_size)), (
                (tx - tx_min) * tile_size,
                (ty - ty_min) * tile_size,
            ))

    full_minx, full_miny, _, _ = tile_bounds_mercator(tx_min, ty_max, z)
    _, _, full_maxx, full_maxy = tile_bounds_mercator(tx_max, ty_min, z)

    def x_to_px(x):
        return int((x - full_minx) / (full_maxx - full_minx) * mosaic_w)

    def y_to_px(y):
        return int((full_maxy - y) / (full_maxy - full_miny) * mosaic_h)

    crop = (
        x_to_px(minx),
        y_to_px(maxy),
        x_to_px(maxx),
        y_to_px(miny),
    )

    cropped = mosaic.crop(crop)

    # 여기서만 최종 크기로 리사이즈
    out = cropped.resize((width, height), Image.Resampling.LANCZOS)

    return out, (minx, miny, maxx, maxy)

# -----------------------------
# GeoJSON (WGS84) — 인터랙티브 지도용
# -----------------------------
def records_to_geojson(records):
    """EPSG:3857 레코드를 WGS84 GeoJSON으로 변환."""
    if not PYPROJ_AVAILABLE:
        return None
    t = get_transformer(3857, 4326)
    features = []
    for rec in records:
        pts = []
        for x, y in rec["points"]:
            lon, lat = transform_xy(x, y, t)
            pts.append([lon, lat])
        if len(pts) < 2:
            continue
        r, g, b = rec["rgb"]
        cat = rec["category"]
        if cat in ("landuse", "boundary") and len(pts) >= 3:
            rings = [pts + [pts[0]]]
            for h in rec.get("holes", []):
                hole_pts = []
                for x, y in h:
                    lon, lat = transform_xy(x, y, t)
                    hole_pts.append([lon, lat])
                if len(hole_pts) >= 4:
                    rings.append(hole_pts + [hole_pts[0]])
            geom = {"type": "Polygon", "coordinates": rings}
        else:
            geom = {"type": "LineString", "coordinates": pts}
        features.append({
            "type": "Feature",
            "properties": {
                "layer": rec["layer"],
                "category": cat,
                "color": "#%02X%02X%02X" % (r, g, b),
            },
            "geometry": geom,
        })
    return {"type": "FeatureCollection", "features": features}


def geojson_center(geojson):
    """GeoJSON에서 대략적인 중심 [lat, lon] 반환."""
    lats, lons = [], []
    for f in geojson.get("features", []):
        coords = f["geometry"].get("coordinates", [])
        if f["geometry"]["type"] == "Polygon":
            coords = coords[0]
        for pt in coords:
            if isinstance(pt[0], (int, float)):
                lons.append(pt[0]); lats.append(pt[1])
    if not lats:
        return [36.5, 127.5]
    return [(min(lats) + max(lats)) / 2, (min(lons) + max(lons)) / 2]


def geojson_bounds(geojson):
    lats, lons = [], []
    for f in geojson.get("features", []):
        coords = f["geometry"].get("coordinates", [])
        if f["geometry"]["type"] == "Polygon":
            coords = coords[0]
        for pt in coords:
            if isinstance(pt[0], (int, float)):
                lons.append(pt[0])
                lats.append(pt[1])
    if not lats:
        return None
    return [[min(lats), min(lons)], [max(lats), max(lons)]]


def show_interactive_map(
    geojson,
    zoom=17,
    boundary_width=4,
    landuse_opacity=0.45,
    line_width=2,
    show_bbox=False,
):
    center = geojson_center(geojson)

    m = folium.Map(
        location=center,
        zoom_start=zoom,
        tiles=None,
        control_scale=True,
        zoom_control=True,
        dragging=True,
        scrollWheelZoom=True,
        doubleClickZoom=True,
        prefer_canvas=True,
    )

    folium.TileLayer(
        tiles=f"https://api.vworld.kr/req/wmts/1.0.0/{VWORLD_API_KEY}/Satellite/{{z}}/{{y}}/{{x}}.jpeg",
        attr="VWorld",
        name="VWorld Satellite",
        overlay=False,
        control=True,
    ).add_to(m)

    def style_boundary(_):
        return {"color": "#FF0000", "weight": boundary_width, "fillOpacity": 0}

    def style_landuse(f):
        color = f["properties"].get("color", "#999999")
        return {"color": color, "weight": 2, "fillColor": color, "fillOpacity": landuse_opacity}

    def style_line(_):
        return {"color": "#111111", "weight": line_width, "fillOpacity": 0}

    layer_defs = [
        ("landuse", "토지이용해치", style_landuse),
        ("line", "계획선/획지선", style_line),
        ("boundary", "구역계", style_boundary),
    ]

    for cat, name, style_func in layer_defs:
        feats = [f for f in geojson["features"] if f["properties"].get("category") == cat]
        if feats:
            folium.GeoJson(
                {"type": "FeatureCollection", "features": feats},
                name=name,
                style_function=style_func,
                tooltip=folium.GeoJsonTooltip(
                    fields=["layer", "category"],
                    aliases=["레이어", "구분"],
                    sticky=True,
                ),
            ).add_to(m)

    bounds = geojson_bounds(geojson)
    if bounds:
        m.fit_bounds(bounds)
        m.options['maxZoom'] = 19
        if show_bbox:
            folium.Rectangle(
                bounds=bounds,
                color="white",
                weight=2,
                fill=False,
                name="추출 영역",
            ).add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)

    # 클릭 시 Leaflet 지도에 생기는 검은 focus 테두리 제거
    m.get_root().html.add_child(folium.Element("""
<style>
.leaflet-container:focus,
.leaflet-container:focus-visible,
.leaflet-interactive:focus,
.leaflet-interactive:focus-visible,
path.leaflet-interactive:focus {
    outline: none !important;
    box-shadow: none !important;
}
.leaflet-control-scale-line {
    background: rgba(255,255,255,0.8);
    padding: 4px;
    font-size: 11px;
}
</style>
"""))
    m.get_root().html.add_child(folium.Element("""
<script>
setTimeout(function() {
    var scales = document.getElementsByClassName('leaflet-control-scale-line');
    for (var i = 0; i < scales.length; i++) {
        if (scales[i].innerHTML.includes('ft')) {
            scales[i].style.display = 'none';
        }
    }
}, 500);
</script>
"""))

    st_folium(m, height=720, use_container_width=True, returned_objects=[])


# -----------------------------
# Rendering
# -----------------------------
def draw_records(
    base_img,
    records,
    bbox3857,
    mode="preview",
    show_boundary=True,
    show_landuse=True,
    show_lines=True,
    boundary_width=4,
    landuse_opacity=145,
    line_width=2,
):
    base_rgba = base_img.convert("RGBA")
    img = base_rgba.copy()
    draw = ImageDraw.Draw(img, "RGBA")
    w, h = img.size
    minx, miny, maxx, maxy = bbox3857

    def to_px(pt):
        x, y = pt
        px = int((x - minx) / (maxx - minx) * w)
        py = int((maxy - y) / (maxy - miny) * h)
        return px, py

    # draw order: landuse fill -> lines -> boundary
    ordered = []
    if show_landuse:
        ordered += [r for r in records if r["category"] == "landuse"]
    if show_lines:
        ordered += [r for r in records if r["category"] == "line"]
    if show_boundary:
        ordered += [r for r in records if r["category"] == "boundary"]

    for rec in ordered:
        pts = [to_px(p) for p in rec["points"]]
        if len(pts) < 2:
            continue
        cat = rec["category"]
        r, g, b = rec["rgb"]
        if cat == "landuse":
            if len(pts) >= 3:
                draw.polygon(pts, fill=(r, g, b, landuse_opacity), outline=(r, g, b, 240))
                for hole in rec.get("holes", []):
                    hole_px = [to_px(p) for p in hole]
                    if len(hole_px) >= 3:
                        mask = Image.new("L", img.size, 0)
                        mask_draw = ImageDraw.Draw(mask)
                        mask_draw.polygon(hole_px, fill=255)
                        if mode == "landuse_hatch":
                            white_bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
                            img.paste(white_bg, (0, 0), mask)
                        else:
                            img.paste(base_rgba, (0, 0), mask)
                        draw = ImageDraw.Draw(img, "RGBA")
                        draw.line(hole_px + [hole_px[0]], fill=(r, g, b, 240), width=max(1, line_width))
        elif cat == "boundary":
            if len(pts) >= 3:
                draw.line(pts + [pts[0]], fill=(255, 0, 0, 255), width=boundary_width)
            else:
                draw.line(pts, fill=(255, 0, 0, 255), width=boundary_width)
        elif cat == "line":
            draw.line(pts, fill=(40, 40, 40, 230), width=line_width)
    return img.convert("RGB")


def make_exports(
    records,
    bbox3857,
    key,
    z,
    width,
    height,
    show_boundary=True,
    show_landuse=True,
    show_lines=True,
    boundary_width=4,
    landuse_opacity=145,
    line_width=2,
):
    sat, padded_bbox = make_satellite_mosaic(bbox3857, key, z=z, width=width, height=height)
    preview = draw_records(
        sat.copy(), records, padded_bbox,
        show_boundary=show_boundary, show_landuse=show_landuse, show_lines=show_lines,
        boundary_width=boundary_width, landuse_opacity=landuse_opacity, line_width=line_width,
    )
    satellite_base = draw_records(
        sat.copy(), records, padded_bbox,
        show_boundary=True, show_landuse=False, show_lines=False,
        boundary_width=boundary_width, landuse_opacity=landuse_opacity, line_width=line_width,
    )
    white = Image.new("RGB", (width, height), (255, 255, 255))
    landuse_hatch = draw_records(
        white, records, padded_bbox,
        show_boundary=True, show_landuse=True, show_lines=True,
        boundary_width=boundary_width, landuse_opacity=255, line_width=line_width,
    )
    return preview, satellite_base, landuse_hatch, padded_bbox


def pil_to_png_bytes(img):
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def build_prompt(rows, site_area):
    lines = [
        "You are given two perfectly aligned images exported from a DXF urban planning drawing:",
        "- Image 1: satellite_base.png = VWorld satellite/base image with exact site boundary.",
        "- Image 2: landuse_hatch.png = land-use hatch map from the same DXF, same bbox and resolution.",
        "",
        "IMPORTANT: DXF metadata below is the authoritative source for layer names, RGB colors, and areas.",
        "Infer land-use programs from the layer names when no explicit program is provided.",
        "Use the images for geometry and visual alignment; use this metadata for semantic interpretation.",
        "",
        f"Total parsed land-use area: approximately {site_area:,.0f} sqm.",
        "",
        "LAND USE LAYER MAPPING:",
    ]
    for r in rows:
        if not r.get("enabled", True):
            continue
        program = r.get("program") or guess_program_from_layer(r["layer"])
        lines.append(f"- Layer '{r['layer']}' / {r.get('rgb','')} / Area ~{r.get('area_sqm',0):,.0f} sqm → {program}")
    lines += [
        "",
        "CORE LOCKS:",
        "- Preserve the site boundary exactly.",
        "- Preserve roads, parcels, land-use hatch geometry, and internal planning lines exactly.",
        "- Do not change or invent geometry outside the site boundary.",
        "- Remove all flat zoning colors in final render; convert each layer into real physical architecture or landscape.",
        "- Roofs must use real systems only: green roofs, solar panels, HVAC, skylights, concrete, metal, or glass.",
        "",
        "OUTPUT TARGET:",
        "A photorealistic Korean urban development aerial masterplan that blends seamlessly with the satellite context.",
    ]
    return "\n".join(lines)



# ──────────────────────────────────────────────────────────────
# DXF 색상 추출
# ──────────────────────────────────────────────────────────────
def aci_to_rgb(aci: int):
    """
    CAD ACI 색상 전체를 RGB로 변환합니다.
    기존 단순 매핑은 ACI 1~9만 처리해서 대부분의 해치가 회색으로 표시되는 문제가 있었습니다.
    """
    try:
        rgb = aci2rgb(int(aci))
        return (int(rgb.r), int(rgb.g), int(rgb.b))
    except Exception:
        aci_map = {
            1: (255, 0, 0),
            2: (255, 255, 0),
            3: (0, 255, 0),
            4: (0, 255, 255),
            5: (0, 0, 255),
            6: (255, 0, 255),
            7: (255, 255, 255),
            8: (128, 128, 128),
            9: (192, 192, 192),
        }
        return aci_map.get(int(aci), (180, 180, 180))


def true_color_to_rgb(true_color: int):
    r = (true_color >> 16) & 255
    g = (true_color >> 8) & 255
    b = true_color & 255
    return (r, g, b)


def guess_preset_from_layer(layer_name: str):
    name = layer_name.lower()
    if "공원" in name or "park" in name:
        return "근린공원·주제공원", "Neighborhood park, tree canopy, walking paths, no buildings"
    if "녹지" in name or "green" in name:
        return "근린공원·주제공원", "Green buffer, dense planting, walking paths, no buildings"
    if "하천" in name or "수변" in name or "water" in name:
        return "하천·수변공간", "River corridor, riparian planting, boardwalk, no buildings"
    if "도로" in name or "road" in name:
        return "광장·공공공지", "Road network, asphalt surface, lane markings, curb lines"
    if "주차" in name or "parking" in name:
        return "주차장", "Surface or structured parking lot, organized layout"
    if "상업" in name or "commercial" in name:
        return "일반상업용지", "Commercial district, mixed retail buildings, active street frontage"
    if "복합" in name or "mixed" in name:
        return "복합용지(혼합개발)", "Mixed-use development zone, podium base with tower elements"
    if "산업" in name or "지원시설" in name or "industrial" in name:
        return "첨단산업단지(R&D·지식산업)", "Business, R&D and light industrial campus, 3~8F"
    if "주거" in name or "res" in name:
        return "공동주택(판상형 아파트)", "Medium-density residential blocks with landscaped courtyards"
    return "[직접입력]", ""


def extract_dxf_layer_colors(dxf_bytes: bytes) -> list:
    if not EZDXF_AVAILABLE:
        return []

    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
        tmp.write(dxf_bytes)
        tmp_path = tmp.name

    try:
        doc = ezdxf.readfile(tmp_path)
        msp = doc.modelspace()

        layer_color_counter = defaultdict(Counter)
        layer_entity_counter = Counter()
        target_types = {"HATCH", "LWPOLYLINE", "POLYLINE"}

        for e in msp:
            try:
                if e.dxftype() not in target_types:
                    continue
                layer_name = e.dxf.layer
                layer_entity_counter[layer_name] += 1
                rgb = None
                if e.dxf.hasattr("true_color") and e.dxf.true_color is not None:
                    rgb = true_color_to_rgb(e.dxf.true_color)
                elif e.dxf.hasattr("color") and int(e.dxf.color) not in [0, 256]:
                    rgb = aci_to_rgb(int(e.dxf.color))
                else:
                    layer = doc.layers.get(layer_name)
                    if layer.dxf.hasattr("true_color") and layer.dxf.true_color is not None:
                        rgb = true_color_to_rgb(layer.dxf.true_color)
                    else:
                        rgb = aci_to_rgb(int(layer.dxf.color))
                layer_color_counter[layer_name][rgb] += 1
            except Exception:
                continue

        rows = []
        for layer_name, counter in layer_color_counter.items():
            if not counter:
                continue
            rgb, _ = counter.most_common(1)[0]
            r, g, b = rgb
            preset, custom_desc = guess_preset_from_layer(layer_name)
            rows.append({
                "layer": layer_name,
                "name": layer_name,
                "r": r, "g": g, "b": b,
                "hex": "#%02X%02X%02X" % (r, g, b),
                "preset": preset,
                "custom_desc": custom_desc,
                "area_sqm": 0.0,
                "tolerance": 20,
                "enabled": True,
                "entity_count": int(layer_entity_counter[layer_name]),
            })
        return rows

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

# ──────────────────────────────────────────────────────────────
# DXF → PNG 렌더링
# ──────────────────────────────────────────────────────────────
def is_landuse_layer(layer_name: str) -> bool:
    n = clean_layer_name(layer_name)
    if is_boundary_layer(layer_name) or is_plan_line_layer(layer_name):
        return False
    keys = [
        "h공원", "h녹지", "h도로", "h도시지원시설", "h보행자전용도로",
        "h복합용지", "h산업", "h산업복합", "h산업지원시설", "h소하천", "h주차장",
        "공원", "녹지", "도로", "도시지원시설", "복합", "산업", "상업", "주차",
        "하천", "수변", "주거", "공동주택", "단독주택", "지원시설", "업무", "상업시설"
    ]
    return any(k in n for k in keys)


def is_boundary_layer(layer_name: str) -> bool:
    n = clean_layer_name(layer_name)
    return "구역계" in n or "boundary" in n or "사업대상지" in n


def dxf_rgb_for_entity(doc, e):
    try:
        if e.dxf.hasattr("true_color") and e.dxf.true_color is not None:
            return true_color_to_rgb(e.dxf.true_color)
        if e.dxf.hasattr("color") and int(e.dxf.color) not in [0, 256]:
            return aci_to_rgb(int(e.dxf.color))
        layer = doc.layers.get(e.dxf.layer)
        if layer.dxf.hasattr("true_color") and layer.dxf.true_color is not None:
            return true_color_to_rgb(layer.dxf.true_color)
        return aci_to_rgb(int(layer.dxf.color))
    except Exception:
        return (180, 180, 180)


def get_dxf_extents(dxf_bytes: bytes):
    if not EZDXF_AVAILABLE:
        return None
    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
        tmp.write(dxf_bytes)
        tmp_path = tmp.name
    try:
        doc = ezdxf.readfile(tmp_path)
        msp = doc.modelspace()
        xs, ys = [], []
        for e in msp:
            try:
                if e.dxftype() == "LWPOLYLINE":
                    for p in e.get_points():
                        xs.append(float(p[0])); ys.append(float(p[1]))
                elif e.dxftype() == "POLYLINE":
                    for v in e.vertices:
                        xs.append(float(v.dxf.location.x)); ys.append(float(v.dxf.location.y))
                elif e.dxftype() == "LINE":
                    xs += [float(e.dxf.start.x), float(e.dxf.end.x)]
                    ys += [float(e.dxf.start.y), float(e.dxf.end.y)]
                elif e.dxftype() == "HATCH":
                    for path in e.paths:
                        if hasattr(path, "vertices"):
                            for v in path.vertices:
                                xs.append(float(v[0])); ys.append(float(v[1]))
            except Exception:
                continue
        if not xs or not ys:
            return None
        return min(xs), min(ys), max(xs), max(ys)
    finally:
        try: os.remove(tmp_path)
        except Exception: pass


def render_dxf_to_png(
    dxf_bytes: bytes,
    mode: str = "preview",
    width: int = 1600,
    height: int = 1200,
    padding_ratio: float = 0.05,
    show_boundary: bool = True,
    show_landuse: bool = True,
    show_plan_lines: bool = True,
    satellite_bytes: bytes = None,
) -> bytes | None:
    """
    mode:
    - preview         : 위성/백판 + 구역계 + 해치 + 선
    - satellite_base  : 위성/백판 + 구역계만
    - landuse_hatch   : 흰 배경 + 구역계 + 토지이용해치
    """
    if not EZDXF_AVAILABLE:
        return None
    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
        tmp.write(dxf_bytes)
        tmp_path = tmp.name
    try:
        doc = ezdxf.readfile(tmp_path)
        msp = doc.modelspace()
        bbox = get_dxf_extents(dxf_bytes)
        if bbox is None:
            return None
        minx, miny, maxx, maxy = bbox
        dx = maxx - minx; dy = maxy - miny
        pad = max(dx, dy) * padding_ratio
        minx -= pad; miny -= pad; maxx += pad; maxy += pad

        img = Image.new("RGB", (width, height), (255, 255, 255))
        if mode in ["preview", "satellite_base"] and satellite_bytes:
            sat = bytes_to_pil(satellite_bytes)
            try:
                sat = sat.resize((width, height), Image.Resampling.LANCZOS)
            except AttributeError:
                sat = sat.resize((width, height), Image.LANCZOS)
            img = sat.copy()

        draw = ImageDraw.Draw(img, "RGBA")

        def xy_to_px(x, y):
            px = int((x - minx) / (maxx - minx) * width)
            py = int((maxy - y) / (maxy - miny) * height)
            return px, py

        def draw_poly(points, stroke, width_px=2, fill=None):
            if len(points) < 2:
                return
            pts = [xy_to_px(x, y) for x, y in points]
            if fill and len(pts) >= 3:
                draw.polygon(pts, fill=fill, outline=stroke)
            else:
                draw.line(pts, fill=stroke, width=width_px, joint="curve")

        for e in msp:
            try:
                layer = e.dxf.layer
                r, g, b = dxf_rgb_for_entity(doc, e)
                bdry = is_boundary_layer(layer)
                lu   = is_landuse_layer(layer)

                if mode == "satellite_base" and not bdry:
                    continue
                if mode == "landuse_hatch" and not (bdry or lu):
                    continue
                if mode == "preview":
                    if bdry and not show_boundary: continue
                    if lu and not show_landuse: continue
                    if (not bdry and not lu) and not show_plan_lines: continue

                stroke = (255, 0, 0, 255) if bdry else (r, g, b, 255)
                width_px = 3 if bdry else 1

                if e.dxftype() == "LWPOLYLINE":
                    pts = [(float(p[0]), float(p[1])) for p in e.get_points()]
                    fill = (r, g, b, 180) if (lu and e.closed) else None
                    draw_poly(pts, stroke, width_px, fill)

                elif e.dxftype() == "POLYLINE":
                    pts = [(float(v.dxf.location.x), float(v.dxf.location.y)) for v in e.vertices]
                    draw_poly(pts, stroke, width_px)

                elif e.dxftype() == "LINE":
                    pts = [(float(e.dxf.start.x), float(e.dxf.start.y)),
                           (float(e.dxf.end.x), float(e.dxf.end.y))]
                    draw_poly(pts, stroke, width_px)

                elif e.dxftype() == "HATCH":
                    for path in e.paths:
                        if hasattr(path, "vertices"):
                            pts = [(float(v[0]), float(v[1])) for v in path.vertices]
                            if len(pts) >= 3:
                                fill = (r, g, b, 190) if lu else None
                                draw_poly(pts, stroke, 1, fill)
            except Exception:
                continue

        return pil_to_png_bytes(img)
    finally:
        try: os.remove(tmp_path)
        except Exception: pass

# ──────────────────────────────────────────────────────────────
# 마스크 추출 + 클립 (경계 이탈 원천 차단)
# ──────────────────────────────────────────────────────────────
def extract_site_mask(plan_bytes: bytes, white_threshold: int = 240) -> "np.ndarray | None":
    """
    흰배경 계획도에서 사이트 마스크 추출.
    반환: 흰색=0(외부), 유색=255(내부) grayscale numpy array
    """
    if not (CV2_AVAILABLE and np is not None):
        return None
    arr = np.array(bytes_to_pil(plan_bytes))
    white = (
        (arr[:, :, 0] >= white_threshold) &
        (arr[:, :, 1] >= white_threshold) &
        (arr[:, :, 2] >= white_threshold)
    )
    mask = np.where(white, 0, 255).astype(np.uint8)
    # 작은 노이즈 제거
    kernel = np.ones((5, 5), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    return mask

def apply_clip(
    generated_bytes: bytes,
    satellite_bytes: bytes,
    site_mask: "np.ndarray",
    feather_radius: int = 3,
) -> bytes:
    """
    생성 결과를 사이트 마스크로 클립.
    마스크 내부 = 생성 픽셀, 외부 = 위성 픽셀.
    """
    if not (CV2_AVAILABLE and np is not None) or site_mask is None:
        return generated_bytes

    gen = np.array(bytes_to_pil(generated_bytes))
    h, w = gen.shape[:2]

    sat = np.array(bytes_to_pil(satellite_bytes))
    sat = cv2.resize(sat, (w, h), interpolation=cv2.INTER_LANCZOS4)

    mask = cv2.resize(site_mask, (w, h), interpolation=cv2.INTER_LINEAR)

    # 엣지 페더링
    if feather_radius > 0:
        mask = cv2.GaussianBlur(mask, (feather_radius * 2 + 1, feather_radius * 2 + 1), 0)

    alpha = mask.astype(np.float32) / 255.0
    alpha3 = np.stack([alpha] * 3, axis=-1)

    result = (gen * alpha3 + sat * (1 - alpha3)).astype(np.uint8)
    return pil_to_png_bytes(Image.fromarray(result))

# ──────────────────────────────────────────────────────────────
# 범례 이미지 — UI 확인용
# ──────────────────────────────────────────────────────────────
def build_legend_image(table: list):
    enabled = [row for row in table if row.get("enabled", True)]
    if not enabled:
        return None

    chip_w, chip_h = 60, 36
    row_h = chip_h + 14
    padding = 16
    text_x = chip_w + padding + 10
    img_w = 560
    header_h = 32
    img_h = header_h + padding * 2 + row_h * len(enabled)

    img = Image.new("RGB", (img_w, img_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, img_w, header_h], fill=(30, 30, 30))
    draw.text((padding, 8), "LAND USE LEGEND", fill=(255, 255, 255))

    for i, row in enumerate(enabled):
        y = header_h + padding + i * row_h
        r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
        draw.rectangle([padding, y, padding + chip_w, y + chip_h],
                       fill=(r, g, b), outline=(80, 80, 80), width=1)
        draw.text((padding, y + chip_h + 1), "R%d G%d B%d" % (r, g, b), fill=(140, 140, 140))

        preset_key = row.get("preset", "")
        custom = row.get("custom_desc", "").strip()
        if custom:
            label_main = custom[:45]
            label_sub = ""
        elif preset_key in ZONE_PRESETS_SIMPLE:
            p = ZONE_PRESETS_SIMPLE[preset_key]
            label_main = p.get("Primary Function", preset_key)[:45]
            label_sub = p.get("prompt_note", "")[:55]
        else:
            label_main = row.get("name", "")[:45]
            label_sub = ""

        draw.text((text_x, y + 4), label_main, fill=(20, 20, 20))
        if label_sub:
            draw.text((text_x, y + 20), label_sub, fill=(90, 90, 90))
        draw.line([padding, y + row_h - 1, img_w - padding, y + row_h - 1],
                  fill=(230, 230, 230), width=1)

    return pil_to_png_bytes(img)

# ──────────────────────────────────────────────────────────────
# 색상 자동 추출
# ──────────────────────────────────────────────────────────────
def extract_dominant_colors(img_bytes: bytes, n_colors: int = 20) -> list:
    if not (CV2_AVAILABLE and np is not None):
        return []
    arr = np.array(bytes_to_pil(img_bytes)).reshape(-1, 3)
    # 흰색 배경만 제외, 검정 도로는 포함
    arr = arr[~((arr[:, 0] > 240) & (arr[:, 1] > 240) & (arr[:, 2] > 240))]
    if len(arr) < 100:
        return []

    arr_q = (arr // 4 * 4).astype(np.int32)
    keys = arr_q[:, 0] * 65536 + arr_q[:, 1] * 256 + arr_q[:, 2]
    unique, counts = np.unique(keys, return_counts=True)
    total = len(arr)

    results = []
    for idx in np.argsort(-counts):
        key = int(unique[idx])
        b = key % 256
        g = (key // 256) % 256
        r = (key // 65536) % 256
        ratio = counts[idx] / total
        if ratio < 0.003:
            continue
        if any(abs(r - er) + abs(g - eg) + abs(b - eb) < 30 for er, eg, eb, _ in results):
            continue
        results.append((r, g, b, float(ratio)))
        if len(results) >= n_colors:
            break
    return results

def build_table_from_detected_colors(
    img_bytes: bytes,
    site_area_sqm: float,
    n_colors: int = 20,
    white_threshold: int = 240,
) -> list:
    """
    흰 배경 토지이용계획도에서 RGB 색상별 면적비를 추정하여
    토지이용 항목 목록을 자동 생성한다.

    기준:
    - 흰색 배경만 제외
    - 검정 도로는 항상 전체면적에 포함
    - 각 픽셀은 가장 가까운 대표 RGB 1개에만 배정
    - 면적 = 사용자 입력 총면적 × 픽셀비율
    """
    if not (CV2_AVAILABLE and np is not None):
        return []

    arr = np.array(bytes_to_pil(img_bytes))

    white_bg = (
        (arr[:, :, 0] >= white_threshold) &
        (arr[:, :, 1] >= white_threshold) &
        (arr[:, :, 2] >= white_threshold)
    )

    valid_mask = ~white_bg
    valid_pixels = arr[valid_mask].astype(np.int16)

    if valid_pixels.shape[0] < 100:
        return []

    colors = extract_dominant_colors(img_bytes, n_colors=n_colors)
    color_list = [(int(r), int(g), int(b)) for r, g, b, _ in colors]

    if not color_list:
        return []

    palette = np.array(color_list, dtype=np.int16)
    diff = valid_pixels[:, None, :] - palette[None, :, :]
    dist2 = np.sum(diff * diff, axis=2)
    nearest_idx = np.argmin(dist2, axis=1)
    counts = np.bincount(nearest_idx, minlength=len(palette))
    total_px = max(1, valid_pixels.shape[0])

    new_rows = []
    for (r, g, b), cnt in zip(color_list, counts):
        ratio = float(cnt) / float(total_px)
        if ratio < 0.003:
            continue
        area_sqm = float(site_area_sqm) * ratio
        is_black = (r < 60 and g < 60 and b < 60)
        new_rows.append({
            "name": "도로" if is_black else "",
            "r": int(r),
            "g": int(g),
            "b": int(b),
            "preset": "[직접입력]" if is_black else "",
            "custom_desc": "Road network, asphalt surface, lane markings, curb lines" if is_black else "",
            "area_sqm": round(area_sqm, 1),
            "tolerance": 20,
            "enabled": True,
        })

    return new_rows

# ──────────────────────────────────────────────────────────────
# RGB 기반 구역 마스크 추출
# ──────────────────────────────────────────────────────────────
def extract_zone_masks(landuse_bytes: bytes, table: list) -> dict:
    if not (CV2_AVAILABLE and np is not None):
        return {}
    rgb_arr = np.array(bytes_to_pil(landuse_bytes))
    h, w = rgb_arr.shape[:2]
    results = {}
    for i, row in enumerate(table):
        if not row.get("enabled", True):
            continue
        r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
        tol = int(row.get("tolerance", 25))
        lo = np.array([max(0, r-tol), max(0, g-tol), max(0, b-tol)], dtype=np.uint8)
        hi = np.array([min(255, r+tol), min(255, g+tol), min(255, b+tol)], dtype=np.uint8)
        mask = cv2.inRange(rgb_arr, lo, hi)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((5, 5), np.uint8), iterations=2)
        area_px = int(np.count_nonzero(mask))
        if area_px < 50:
            continue
        m = cv2.moments(mask)
        cx = int(m["m10"] / m["m00"]) if m["m00"] > 0 else w // 2
        cy = int(m["m01"] / m["m00"]) if m["m00"] > 0 else h // 2
        cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if cnts:
            x_, y_, ww, hh = cv2.boundingRect(np.concatenate(cnts, axis=0))
        else:
            x_, y_, ww, hh = 0, 0, w, h
        results[i] = {
            "mask": mask, "area_px": area_px,
            "centroid": (cx, cy), "bbox": (x_, y_, ww, hh),
            "img_size": (w, h),
        }
    return results

def describe_position(cx, cy, w, h) -> str:
    lr = "western" if cx < w * 0.4 else ("eastern" if cx > w * 0.6 else "central")
    tb = "northern" if cy < h * 0.4 else ("southern" if cy > h * 0.6 else "central")
    if lr == "central" and tb == "central":
        return "center of the site"
    if lr == "central":
        return "%s part of the site" % tb
    if tb == "central":
        return "%s part of the site" % lr
    return "%s-%s part of the site" % (tb, lr)

# ──────────────────────────────────────────────────────────────
# 공원/녹지 여부 판별
# ──────────────────────────────────────────────────────────────
def should_keep_color(row: dict) -> bool:
    preset = row.get("preset", "")
    custom = row.get("custom_desc", "").strip().lower()

    no_building_presets = {"근린공원·주제공원", "하천·수변공간", "광장·공공공지"}
    if preset in no_building_presets:
        return True

    no_building_keywords = [
        "no buildings", "park", "trail", "forest", "golf",
        "pool", "water", "green buffer", "landscape", "garden",
        "walking", "pedestrian", "plaza", "retention", "healing",
        "open space", "lawn", "wetland",
    ]
    return any(k in custom for k in no_building_keywords)

# ──────────────────────────────────────────────────────────────
# 합성 입력 이미지 생성
# 위성(배경) + site 내부 흰색 + 검정 경계선 + [Z] 레이블
# 공원/녹지 구역은 원본 색상 유지
# ──────────────────────────────────────────────────────────────
def build_composite_with_labels(
    landuse_bytes: bytes, table: list, sat_bytes: bytes = None
) -> tuple:
    if not (CV2_AVAILABLE and np is not None):
        return landuse_bytes, {}

    landuse_arr = np.array(bytes_to_pil(landuse_bytes))
    h, w = landuse_arr.shape[:2]

    # 배경 준비
    if sat_bytes:
        sat = bytes_to_pil(sat_bytes)
        try:
            sat = sat.resize((w, h), Image.Resampling.LANCZOS)
        except AttributeError:
            sat = sat.resize((w, h), Image.LANCZOS)
        result = np.array(sat).copy()
    else:
        result = np.ones((h, w, 3), dtype=np.uint8) * 255

    # 검정 경계선 마스크
    gray = cv2.cvtColor(landuse_arr, cv2.COLOR_RGB2GRAY)
    _, black_mask = cv2.threshold(gray, 60, 255, cv2.THRESH_BINARY_INV)

    # 흰색 외부 배경 마스크
    white_bg = (
        (landuse_arr[:, :, 0] > 240) &
        (landuse_arr[:, :, 1] > 240) &
        (landuse_arr[:, :, 2] > 240)
    )

    # site 내부 → 흰색
    site_interior = ~white_bg & (black_mask == 0)
    result[site_interior] = [255, 255, 255]

    # 검정 경계선 복원
    result[black_mask > 0] = [30, 30, 30]

    # Z번호 그룹화
    def get_zone_key(row: dict) -> str:
        preset = row.get("preset", "[직접입력]")
        custom = row.get("custom_desc", "").strip()
        if custom:
            return "custom::" + custom[:40]
        return preset

    zone_key_to_z = {}
    zone_label_map = {}
    z_counter = [1]

    def get_z_num(row):
        key = get_zone_key(row)
        if key not in zone_key_to_z:
            zone_key_to_z[key] = z_counter[0]
            zone_label_map["Z%d" % z_counter[0]] = row
            z_counter[0] += 1
        return zone_key_to_z[key]

    # 구역별 처리
    for i, row in enumerate(table):
        if not row.get("enabled", True):
            continue
        r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
        tol = int(row.get("tolerance", 25))
        lo = np.array([max(0, r-tol), max(0, g-tol), max(0, b-tol)], dtype=np.uint8)
        hi = np.array([min(255, r+tol), min(255, g+tol), min(255, b+tol)], dtype=np.uint8)
        mask = cv2.inRange(landuse_arr, lo, hi)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((5, 5), np.uint8), iterations=2)
        if np.count_nonzero(mask) < 200:
            continue

        if should_keep_color(row):
            result[mask > 0] = [r, g, b]
            continue

        z_num = get_z_num(row)
        label_text = "[Z%d]" % z_num

        num_comp, _, stats, centroids = cv2.connectedComponentsWithStats(mask, connectivity=8)
        for ci in range(1, num_comp):
            comp_area = stats[ci, cv2.CC_STAT_AREA]
            if comp_area < 100:
                continue
            cx = int(centroids[ci][0])
            cy = int(centroids[ci][1])

            font_scale = max(0.4, min(1.5, comp_area ** 0.5 / 130))
            thickness = max(1, int(font_scale * 2))
            (tw, th), _ = cv2.getTextSize(
                label_text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness
            )
            tx = max(tw // 2, min(w - tw, cx - tw // 2))
            ty = max(th, min(h - 4, cy + th // 2))

            cv2.rectangle(result,
                (tx - 2, ty - th - 2), (tx + tw + 2, ty + 2),
                (255, 255, 255), -1)
            cv2.putText(result, label_text, (tx, ty),
                cv2.FONT_HERSHEY_SIMPLEX, font_scale,
                (20, 20, 20), thickness, cv2.LINE_AA)

    return pil_to_png_bytes(Image.fromarray(result)), zone_label_map

# ──────────────────────────────────────────────────────────────
# 흰색 경계선 제거 후처리
# ──────────────────────────────────────────────────────────────
def remove_white_lines(img_bytes: bytes) -> bytes:
    if not (CV2_AVAILABLE and np is not None):
        return img_bytes
    try:
        arr = np.array(bytes_to_pil(img_bytes))
        bgr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
        _, white = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY)
        num, lbl, stats, _ = cv2.connectedComponentsWithStats(white, connectivity=8)
        line_mask = np.zeros_like(white)
        for i in range(1, num):
            area = stats[i, cv2.CC_STAT_AREA]
            bw = stats[i, cv2.CC_STAT_WIDTH]
            bh = stats[i, cv2.CC_STAT_HEIGHT]
            if area < 5000 and (bw < 15 or bh < 15):
                line_mask[lbl == i] = 255
        if np.sum(line_mask) < 10:
            return img_bytes
        dilated = cv2.dilate(line_mask, np.ones((3, 3), np.uint8), iterations=1)
        result = cv2.inpaint(bgr, dilated, 4, cv2.INPAINT_TELEA)
        return pil_to_png_bytes(Image.fromarray(cv2.cvtColor(result, cv2.COLOR_BGR2RGB)))
    except Exception:
        return img_bytes

# ──────────────────────────────────────────────────────────────
# PASS1 프롬프트
# ──────────────────────────────────────────────────────────────
def build_pass1_prompt(
    table: list, zone_masks: dict, site_area: float, zone_label_map: dict = None
) -> str:
    lines = [
        "You are given TWO images exported from a DXF urban planning drawing:",
        "- Image 1: satellite_base.png — satellite/base canvas with the exact site boundary.",
        "- Image 2: landuse_hatch.png — land-use hatch map exported from the same DXF,",
        "  perfectly aligned with Image 1.",
        "",
        "The DXF metadata listed below is the AUTHORITATIVE SOURCE for layer names, RGB colors,",
        "areas, and land-use programs. Use the images only for geometry and visual alignment.",
        "Do NOT infer land-use type from image color alone — always follow DXF metadata.",
        "",
        "TASK: Convert the land-use hatch plan into a premium top-down 2D urban masterplan illustration.",
        "The result must read as a single unified aerial image consistent with Image 1.",
        "",
        "MATERIALS — EVERY SURFACE MUST USE REAL PHYSICAL MATERIALS ONLY:",
        "Rooftops: green roof vegetation, solar panel arrays, HVAC units,",
        "  skylights, exposed concrete, or reflective glass.",
        "Facades: glass curtain wall, brick, exposed concrete, metal cladding,",
        "  stone panel, ceramic tile, painted plaster.",
        "Ground: asphalt, concrete pavement, gravel, natural soil,",
        "  grass, tree canopy, stone paving, pedestrian tiles.",
        "No surface may display a flat zone color — any flat-colored surface",
        "is a generation error. Override with permitted materials above.",
        "",
        "ZONE RULES:",
        "- White [Z] zones: fill fully with buildings, roads, landscape per zone type below.",
        "- Colored zones: enhance with trees, texture, paths. NO buildings whatsoever.",
        "- Preserve ALL zone boundaries exactly. Do NOT redraw or merge zones.",
        "- NO text, labels, or zone numbers in the output.",
        "- TOTAL SITE AREA: ~%s sqm." % "{:,.0f}".format(site_area),
        "",
        "BUILDING ZONE TYPES:",
    ]

    if zone_label_map:
        for z_key, row in zone_label_map.items():
            preset_key = row.get("preset", "[직접입력]")
            custom = row.get("custom_desc", "").strip()
            if custom:
                desc = custom
            elif preset_key in ZONE_PRESETS_SIMPLE:
                desc = ZONE_PRESETS_SIMPLE[preset_key].get("prompt_note", preset_key)
            else:
                desc = row.get("name", z_key)
            user_area = row.get("area_sqm", 0)
            area_str = " | ~%s sqm" % "{:,.0f}".format(user_area) if user_area > 0 else ""
            layer_name = row.get("layer", "")
            rgb_text = "RGB(%d,%d,%d)" % (int(row["r"]), int(row["g"]), int(row["b"]))
            if layer_name:
                lines.append("  [%s] Layer '%s' / %s = %s%s" % (z_key, layer_name, rgb_text, desc, area_str))
            else:
                lines.append("  [%s] = %s%s" % (z_key, desc, area_str))
    else:
        seen_rgb = set()
        for row in table:
            if not row.get("enabled", True):
                continue
            r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
            if (r, g, b) in seen_rgb:
                continue
            seen_rgb.add((r, g, b))
            preset_key = row.get("preset", "[직접입력]")
            custom = row.get("custom_desc", "").strip()
            desc = custom if custom else ZONE_PRESETS_SIMPLE.get(preset_key, {}).get("prompt_note", row.get("name", ""))
            layer_name = row.get("layer", "")
            if layer_name:
                lines.append("  Layer '%s' / RGB(%d,%d,%d) = %s" % (layer_name, r, g, b, desc))
            else:
                lines.append("  RGB(%d,%d,%d) = %s" % (r, g, b, desc))

    lines += [
        "",
        "ARCHITECTURE:",
        "Varied massing per zone: stepped volumes, courtyard typologies, podium-tower compositions.",
        "Allow 1–2 landmark buildings within their respective footprints.",
        "All facades: realistic window systems, material transitions, floor-level articulation.",
        "",
        "OUTPUT STYLE — PREMIUM 2D MASTERPLAN:",
        "Style: high-end Korean urban development competition board.",
        "Color palette:",
        "  Residential: warm beige/cream buildings, grey slate rooftops, soft drop shadows.",
        "  Parks/green: rich dark green canopy over light green lawn, circular tree shadows.",
        "  Public: terracotta/orange accent roofs, civic plazas.",
        "  Water: vivid blue-green with subtle ripple texture.",
        "  Roads: light warm grey asphalt, white lane markings, curb lines.",
        "Buildings: many varied footprints — L-shape, U-shape, courtyard, slab, point tower.",
        "Landscape: lush dark green canopies, light green lawns, dense street trees.",
        "Quality: competition-board quality. Rich, dense, colorful, highly detailed.",
        "Crisp clean edges. No text, no labels, no zone markers in output.",
        "",
        "HARD CONSTRAINTS:",
        "DO NOT generate anything in white areas outside the site boundary.",
        "DO NOT produce CG, cartoon, or plastic-looking output.",
    ]
    return "\n".join(lines).strip()

# ──────────────────────────────────────────────────────────────
# PASS2 프롬프트
# ──────────────────────────────────────────────────────────────
FIXED_QUALITY_OBLIQUE = (
    "RENDER QUALITY — PHOTOREALISTIC (competition-grade archviz):\n"
    "Lighting: Late-afternoon golden-hour sun, low angle. Strong directional shadows. "
    "Warm orange-gold on sunlit faces. Cool blue-grey in shadows. Filmic HDR.\n"
    "Glass: Fresnel reflections, sky color reflected, sharp sun glints on curtain walls.\n"
    "Brick/concrete: Visible mortar joints, surface grain, shadow lines under overhangs.\n"
    "Vegetation: 3D volumetric tree canopies, translucent leaf edges, cast shadows.\n"
    "Roads: Asphalt texture, curb shadow lines, sidewalk paving patterns.\n"
    "No blur, no cartoon shading, no flat uniform colors. Photorealistic throughout.\n"
)

def build_pass2_prompt(table: list) -> str:
    lines = [
        "Convert the input 2D masterplan into a photorealistic 3D oblique aerial archviz render.",
        "",
        "MATERIALS — EVERY SURFACE MUST USE REAL PHYSICAL MATERIALS ONLY:",
        "Rooftops: green roof vegetation, solar panel arrays, HVAC units,",
        "  skylights, exposed concrete, or reflective glass.",
        "Facades: glass curtain wall, brick, exposed concrete, metal cladding,",
        "  stone panel, ceramic tile, painted plaster.",
        "Ground: asphalt, concrete pavement, gravel, natural soil,",
        "  grass, tree canopy, stone paving, pedestrian tiles.",
        "No surface may display a flat zone color — any flat-colored surface",
        "is a generation error. Override with permitted materials above.",
        "",
        "LAND USE REFERENCE:",
    ]

    seen_rgb = set()
    for row in table:
        if not row.get("enabled", True):
            continue
        r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
        if (r, g, b) in seen_rgb:
            continue
        seen_rgb.add((r, g, b))
        preset_key = row.get("preset", "[직접입력]")
        custom = row.get("custom_desc", "").strip()
        desc = custom[:80] if custom else ZONE_PRESETS_SIMPLE.get(preset_key, {}).get("prompt_note", row.get("name", ""))
        layer_name = row.get("layer", "")
        if layer_name:
            lines.append("  Layer '%s' / RGB(%d,%d,%d) = %s" % (layer_name, r, g, b, desc))
        else:
            lines.append("  RGB(%d,%d,%d) = %s" % (r, g, b, desc))

    lines += [
        "",
        "Keep all geometry exactly as in the input. Do NOT redesign or crop.",
        "CAMERA: 45–55 degree oblique aerial view, matched to input perspective.",
        "",
        FIXED_QUALITY_OBLIQUE,
        "HARD CONSTRAINTS:",
        "DO NOT produce CG, cartoon, or plastic-looking output.",
        "DO NOT render any text or labels.",
    ]
    return "\n".join(lines).strip()

# ──────────────────────────────────────────────────────────────
# PASS3 프롬프트
# ──────────────────────────────────────────────────────────────
def build_pass3_prompt(angle: int) -> str:
    return (
        "Keep everything exactly the same. "
        "Only change the camera angle to %d degrees oblique aerial view." % angle
    )

# ──────────────────────────────────────────────────────────────
# Gemini helpers
# ──────────────────────────────────────────────────────────────
def _part_from_text(text):
    if GENAI_TYPES_AVAILABLE:
        try: return types.Part.from_text(text=text)
        except: return types.Part(text=text)
    return {"text": text}

def _part_from_bytes(data, mime="image/png"):
    if GENAI_TYPES_AVAILABLE:
        try: return types.Part.from_bytes(data=data, mime_type=mime)
        except: return types.Part(inline_data=types.Blob(mime_type=mime, data=data))
    return {"inline_data": {"mime_type": mime, "data": data}}

def make_contents(prompt: str, images: list) -> list:
    return [_part_from_text(prompt)] + [_part_from_bytes(b) for b in images]

def get_image_from_resp(resp):
    parts = getattr(resp, "parts", None)
    if parts is None:
        cands = getattr(resp, "candidates", [])
        content = getattr(cands[0], "content", None) if cands else None
        parts = getattr(content, "parts", []) if content else []
    for part in parts:
        inline = getattr(part, "inline_data", None)
        if inline:
            data = getattr(inline, "data", None)
            if data:
                return data
    return None

# ──────────────────────────────────────────────────────────────
# 이미지 선택 UI
# ──────────────────────────────────────────────────────────────
def render_selector(outputs, sel_key, label):
    if not outputs:
        return None
    for row_start in range(0, len(outputs), 2):
        cols = st.columns(2, gap="medium")
        for j, img_bytes in enumerate(outputs[row_start:row_start+2]):
            i = row_start + j
            with cols[j]:
                im = bytes_to_pil(img_bytes)
                w, h = im.size
                is_sel = (st.session_state[sel_key] == i)
                border = "2px solid #2563EB" if is_sel else "1px solid #E5E7EB"
                st.markdown('<div style="border:%s;border-radius:10px;padding:6px;">' % border,
                            unsafe_allow_html=True)
                st.image(im, caption="%s #%d — %dx%d" % (label, i+1, w, h),
                         use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)
                btn_label = "선택됨 #%d" % (i+1) if is_sel else "선택 #%d" % (i+1)
                if st.button(btn_label, key="%s_btn_%d" % (sel_key, i),
                             use_container_width=True,
                             type="primary" if is_sel else "secondary"):
                    st.session_state[sel_key] = i
                    st.rerun()
    return outputs[st.session_state[sel_key]]


# ──────────────────────────────────────────────────────────────
# DXF 추출 결과 → 기존 토지이용표 연결 유틸
# ──────────────────────────────────────────────────────────────
def parse_rgb_text(rgb_text: str):
    m = re.search(r"RGB\((\d+),\s*(\d+),\s*(\d+)\)", str(rgb_text))
    if not m:
        return 180, 180, 180
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def dxf_rows_to_landuse_table(rows: list) -> list:
    out = []
    for row in rows:
        layer = str(row.get("layer", "")).strip()
        r, g, b = parse_rgb_text(row.get("rgb", ""))
        preset, custom_desc = guess_preset_from_layer(layer)
        area = float(row.get("area_sqm", 0.0) or 0.0)
        out.append({
            "layer": layer,
            "name": layer,
            "r": r,
            "g": g,
            "b": b,
            "hex": "#%02X%02X%02X" % (r, g, b),
            "preset": preset,
            "custom_desc": custom_desc,
            "area_sqm": area,
            "tolerance": 20,
            "enabled": bool(row.get("enabled", True)),
            "entity_count": int(row.get("entity_count", 0) or 0),
        })
    return out


def reset_generated_images():
    st.session_state.pass1_outputs = []
    st.session_state.pass1_selected_idx = 0
    st.session_state.pass1_output_bytes = None
    st.session_state.pass2_outputs = []
    st.session_state.pass2_selected_idx = 0
    st.session_state.pass2_output_bytes = None
    st.session_state.pass3_outputs = []
    st.session_state.pass3_selected_idx = 0

# ──────────────────────────────────────────────────────────────
# 상단 네비게이션
# ──────────────────────────────────────────────────────────────
STEPS = ["① DXF 입력·PNG 생성", "② 레이어·프롬프트 확인", "③ 조감도 생성"]
cur_step = st.session_state.step

cols_nav = st.columns(len(STEPS))
for _i, _label in enumerate(STEPS):
    with cols_nav[_i]:
        _active = (_i == cur_step)
        _done = (_i < cur_step)
        _bg = "#2563EB" if _active else ("#DCFCE7" if _done else "#F9FAFB")
        _fg = "#fff" if _active else ("#15803D" if _done else "#9CA3AF")
        _border = "#2563EB" if _active else ("#86EFAC" if _done else "#E5E7EB")
        st.markdown(
            '<div style="background:%s;color:%s;border:2px solid %s;'
            'border-radius:10px;padding:10px 0;text-align:center;'
            'font-weight:700;font-size:14px;">%s%s</div>'
            % (_bg, _fg, _border, "✓ " if _done else "", _label),
            unsafe_allow_html=True
        )

nav_c1, nav_c2, _ = st.columns([1, 1, 4])
with nav_c1:
    if st.button("◀ 이전", disabled=(cur_step == 0)):
        st.session_state.step -= 1
        st.rerun()
with nav_c2:
    _can_next = (
        (cur_step == 0 and st.session_state.dxf_landuse_hatch_bytes is not None) or
        (cur_step == 1)
    )
    if st.button("다음 ▶", disabled=(cur_step >= 2 or not _can_next)):
        st.session_state.step += 1
        st.rerun()

st.markdown(
    '<div style="height:3px;background:linear-gradient(90deg,#2563EB,#1E293B 80%);'
    'border-radius:2px;margin:16px 0 28px 0;"></div>',
    unsafe_allow_html=True
)

# 에러 표시
if st.session_state._errors:
    for err in st.session_state._errors:
        st.error(err)
    st.session_state._errors = []

# ══════════════════════════════════════════════════════════════
# STEP 0: DXF 입력 + VWorld 기반 PNG 2장 생성
# ══════════════════════════════════════════════════════════════
if cur_step == 0:
    st.markdown('<div class="section-header">① DXF 입력·PNG 생성</div>', unsafe_allow_html=True)
    st.caption(
        "DXF를 업로드하면 VWorld 위성백판 위에서 구역계·토지이용해치를 확인하고, "
        "조감도 생성에 필요한 satellite_base.png와 landuse_hatch.png 2장을 같은 화면에서 생성합니다."
    )

    missing = []
    if not EZDXF_AVAILABLE:
        missing.append("ezdxf")
    if not PYPROJ_AVAILABLE:
        missing.append("pyproj")
    if not REQUESTS_AVAILABLE:
        missing.append("requests")
    if st_folium is None:
        missing.append("streamlit-folium")
    if folium is None:
        missing.append("folium")

    if missing:
        st.error("필수 패키지가 없습니다: " + ", ".join(missing))
        st.code("pip install ezdxf pyproj requests folium streamlit-folium google-genai opencv-python openpyxl", language="bash")
        st.stop()

    dxf_file = st.file_uploader(
        "DXF 업로드 — 구역계/토지이용해치/계획선/획지선 포함",
        type=["dxf"],
        key="up_dxf_main"
    )

    c_opt1, c_opt2 = st.columns([1.3, 1])
    with c_opt1:
        crs_label = st.selectbox(
            "DXF 좌표계",
            list(CRS_OPTIONS.keys()),
            index=0,
            help="대부분 구 한국측지계 CAD는 EPSG:5174인 경우가 많습니다."
        )
        src_epsg = CRS_OPTIONS[crs_label]
    with c_opt2:
        img_size_label = st.selectbox("PNG 출력 크기", list(IMG_SIZE_OPTIONS.keys()), index=2)
        out_w, out_h = IMG_SIZE_OPTIONS[img_size_label]

    export_zoom = st.slider(
        "VWorld 위성 타일 줌 레벨",
        min_value=16,
        max_value=20,
        value=DEFAULT_EXPORT_ZOOM,
        step=1,
        help="19 권장. 너무 높이면 일부 지역에서 타일 누락 또는 속도 저하가 있을 수 있습니다."
    )

    st.markdown('<div class="sub-label">계획부지 전체 면적</div>', unsafe_allow_html=True)
    st.session_state.site_area_sqm = st.number_input(
        "계획부지 전체면적 (㎡)",
        min_value=1.0,
        value=float(st.session_state.site_area_sqm),
        step=1000.0,
        format="%.0f",
    )

    if dxf_file:
        raw = dxf_file.getvalue()
        if raw != st.session_state.dxf_bytes:
            st.session_state.dxf_bytes = raw
            st.session_state["_auto_generated"] = False
            st.session_state.dxf_rows = []
            st.session_state.dxf_layer_table = []
            st.session_state.land_use_table = []
            st.session_state.records = []
            st.session_state.dxf_geojson = None
            st.session_state.gdal_geojson = None
            st.session_state.bbox_3857 = None
            st.session_state.preview_png = None
            st.session_state.satellite_base_png = None
            st.session_state.landuse_hatch_png = None
            st.session_state.dxf_satellite_base_bytes = None
            st.session_state.dxf_landuse_hatch_bytes = None
            st.session_state.img_sat_bytes = None
            st.session_state.img_landuse_bytes = None
            reset_generated_images()

            try:
                with st.spinner("DXF를 GIS geometry로 변환하고 레이어/RGB/면적을 계산하는 중..."):
                    gj = dxf_to_geojson_by_gdal(raw, src_epsg)
                    records, bbox = records_from_gdal_geojson(gj, raw, src_epsg)
                    area_map = compute_hatch_area_by_layer(raw)
                    rows = build_dxf_layer_table(records, area_map=area_map)

                if not records or bbox is None:
                    st.error("DXF에서 구역계/토지이용해치/계획선 geometry를 읽지 못했습니다. 레이어명과 좌표계를 확인하세요.")
                else:
                    st.session_state.gdal_geojson = gj
                    st.session_state.records = records
                    st.session_state.bbox_3857 = bbox
                    st.session_state.dxf_rows = rows
                    st.session_state.dxf_geojson = records_to_geojson(records)
                    st.session_state.dxf_layer_table = dxf_rows_to_landuse_table(rows)

                    total_area = sum(float(r.get("area_sqm", 0.0) or 0.0) for r in rows if r.get("enabled", True))
                    if total_area > 0:
                        st.session_state.site_area_sqm = total_area

                    st.session_state.land_use_table = [
                        {
                            "layer": r["layer"],
                            "name": r["name"],
                            "r": r["r"], "g": r["g"], "b": r["b"],
                            "preset": r["preset"],
                            "custom_desc": r["custom_desc"],
                            "area_sqm": r.get("area_sqm", 0.0),
                            "tolerance": r.get("tolerance", 20),
                            "enabled": r.get("enabled", True),
                        }
                        for r in st.session_state.dxf_layer_table
                    ]
                    st.session_state["_auto_generated"] = True
                    st.session_state.prompt_text = build_prompt(rows, total_area)
                    st.success("DXF 파싱 완료. 아래에서 지도 확인 후 PNG 2장을 생성하세요.")
            except Exception as e:
                st.error("DXF 변환 오류: %s" % str(e))

    if st.session_state.dxf_geojson:
        st.markdown('<div class="sub-label">인터랙티브 지도 확인</div>', unsafe_allow_html=True)
        mc1, mc2, mc3 = st.columns(3)
        boundary_width = mc1.slider("구역계 두께", 1, 8, 4)
        landuse_opacity = mc2.slider("토지이용해치 투명도", 0.05, 0.90, 0.45, 0.05)
        line_width = mc3.slider("계획선/획지선 두께", 1, 5, 2)

        show_interactive_map(
            st.session_state.dxf_geojson,
            zoom=DEFAULT_MAP_ZOOM,
            boundary_width=boundary_width,
            landuse_opacity=landuse_opacity,
            line_width=line_width,
        )

        if st.session_state.dxf_layer_table:
            st.markdown('<div class="sub-label">추출된 레이어·RGB·면적</div>', unsafe_allow_html=True)
            st.dataframe(
                [
                    {
                        "사용": r.get("enabled", True),
                        "레이어": r["layer"],
                        "RGB": "%d,%d,%d" % (r["r"], r["g"], r["b"]),
                        "HEX": r.get("hex", ""),
                        "면적(㎡)": r.get("area_sqm", 0.0),
                        "추정 프리셋": r.get("preset", ""),
                    }
                    for r in st.session_state.dxf_layer_table
                ],
                use_container_width=True,
            )

        st.markdown("---")
        st.markdown("### 조감도 입력용 PNG 2장 생성")
        st.caption("satellite_base: VWorld 위성백판 + 구역계 / landuse_hatch: 흰 백판 + 토지이용해치 + 계획선")

        if st.button("PNG 2장 생성", type="primary", use_container_width=True):
            try:
                with st.spinner("VWorld 위성 타일을 받아 PNG 2장을 생성하는 중..."):
                    preview, sat_base, landuse_hatch, padded_bbox = make_exports(
                        st.session_state.records,
                        st.session_state.bbox_3857,
                        VWORLD_API_KEY,
                        export_zoom,
                        out_w,
                        out_h,
                        show_boundary=True,
                        show_landuse=True,
                        show_lines=True,
                        boundary_width=boundary_width,
                        landuse_opacity=int(255 * landuse_opacity),
                        line_width=line_width,
                    )

                st.session_state.preview_png = pil_to_png_bytes(preview)
                st.session_state.satellite_base_png = pil_to_png_bytes(sat_base)
                st.session_state.landuse_hatch_png = pil_to_png_bytes(landuse_hatch)
                st.session_state.dxf_satellite_base_bytes = st.session_state.satellite_base_png
                st.session_state.dxf_landuse_hatch_bytes = st.session_state.landuse_hatch_png
                st.session_state.img_sat_bytes = st.session_state.satellite_base_png
                st.session_state.img_landuse_bytes = st.session_state.landuse_hatch_png
                st.session_state.export_bbox_3857 = padded_bbox
                reset_generated_images()
                st.success("PNG 2장 생성 완료. 바로 다음 단계로 이동할 수 있습니다.")
                st.rerun()
            except Exception as e:
                st.error("PNG 생성 오류: %s" % str(e))

    if st.session_state.dxf_satellite_base_bytes and st.session_state.dxf_landuse_hatch_bytes:
        pa, pb = st.columns(2)
        with pa:
            st.image(
                bytes_to_pil(st.session_state.dxf_satellite_base_bytes),
                caption="Image 1: satellite_base.png",
                use_container_width=True
            )
            st.download_button(
                "⬇️ satellite_base.png",
                data=st.session_state.dxf_satellite_base_bytes,
                file_name="satellite_base.png",
                mime="image/png",
                use_container_width=True,
            )
        with pb:
            st.image(
                bytes_to_pil(st.session_state.dxf_landuse_hatch_bytes),
                caption="Image 2: landuse_hatch.png",
                use_container_width=True
            )
            st.download_button(
                "⬇️ landuse_hatch.png",
                data=st.session_state.dxf_landuse_hatch_bytes,
                file_name="landuse_hatch.png",
                mime="image/png",
                use_container_width=True,
            )

        with st.expander("DXF 메타데이터 기반 프롬프트 초안", expanded=False):
            st.code(st.session_state.prompt_text or "", language="text")

        st.success("'다음 ▶'으로 이동하면 이 PNG 2장과 레이어 메타데이터가 그대로 조감도 생성 단계에 연결됩니다.")
    else:
        st.warning("DXF 업로드 후 'PNG 2장 생성'을 눌러야 다음 단계로 이동할 수 있습니다.")


# ══════════════════════════════════════════════════════════════
# STEP 1: 토지이용계획표
# ══════════════════════════════════════════════════════════════
elif cur_step == 1:
    st.markdown('<div class="section-header">② 레이어·프롬프트 확인</div>', unsafe_allow_html=True)
    st.caption("STEP ①에서 파싱된 레이어명, RGB, 면적, 용도 추정값을 확인·수정합니다.")

    if st.button("색상/면적 다시 계산", type="secondary"):
        st.session_state["_auto_generated"] = False
        st.session_state.land_use_table = []
        st.rerun()

    # DXF 레이어 기반 자동 생성
    with st.expander("🧩 DXF 레이어 기반 토지이용표 자동 생성", expanded=True):
        st.caption("DXF에서 추출한 레이어명/RGB를 토지이용계획표 초안으로 적용합니다.")
        if st.session_state.dxf_layer_table:
            st.dataframe(
                [
                    {
                        "레이어": r["layer"],
                        "RGB": "%d,%d,%d" % (r["r"], r["g"], r["b"]),
                        "HEX": r["hex"],
                        "추정 프리셋": r["preset"],
                        "객체수": r["entity_count"],
                    }
                    for r in st.session_state.dxf_layer_table
                ],
                use_container_width=True,
            )
            if st.button("DXF 레이어/RGB를 토지이용표에 적용", type="primary"):
                st.session_state.land_use_table = [
                    {
                        "layer": r["layer"],
                        "name": r["name"],
                        "r": r["r"], "g": r["g"], "b": r["b"],
                        "preset": r["preset"],
                        "custom_desc": r["custom_desc"],
                        "area_sqm": r.get("area_sqm", 0.0),
                        "tolerance": r.get("tolerance", 20),
                        "enabled": r.get("enabled", True),
                    }
                    for r in st.session_state.dxf_layer_table
                ]
                st.session_state["_auto_generated"] = True
                st.success("DXF 기반 토지이용표가 적용되었습니다.")
                st.rerun()
        else:
            st.info("STEP ①에서 DXF를 먼저 업로드하세요.")

    # 엑셀 업로드
    with st.expander("📥 엑셀로 토지이용계획표 불러오기", expanded=False):
        st.caption("컬럼 순서: 용도명 / R / G / B / 면적(㎡) / 용도설명(영문) / 프리셋(선택)")

        if st.button("📄 템플릿 다운로드"):
            import io as _io
            try:
                import openpyxl as _xl
                _wb = _xl.Workbook()
                _ws = _wb.active
                _ws.title = "토지이용계획표"
                _ws.append(["용도명", "R", "G", "B", "면적(㎡)", "용도설명(영문)", "프리셋"])
                for _r in [
                    ("단독주택",  255, 255, 127, 10000, "Low-rise detached housing, 2~3F, garden plots", "단독주택"),
                    ("공원",        0, 165,   0, 15000, "Neighborhood park, tree canopy, walking paths", "근린공원·주제공원"),
                    ("숙박시설",  255, 191, 127, 12000, "resort hotel with amenity facilities, 5~10F, warm facade", "[직접입력]"),
                    ("치유의숲",  127, 255,   0, 20000, "healing forest with walking trails, meditation zones, no buildings", "[직접입력]"),
                ]:
                    _ws.append(_r)
                _buf = _io.BytesIO()
                _wb.save(_buf)
                _buf.seek(0)
                st.download_button(
                    "⬇️ 템플릿 xlsx 다운로드", data=_buf.getvalue(),
                    file_name="토지이용계획표_템플릿.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ImportError:
                st.error("openpyxl 패키지가 필요합니다.")

        xl_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"], key="xl_upload")
        if xl_file and st.button("테이블에 적용", type="primary", key="apply_xl"):
            try:
                import openpyxl as _xl
                import io as _io
                _wb = _xl.load_workbook(_io.BytesIO(xl_file.getvalue()), data_only=True)
                _ws = _wb.active
                _new_rows = []
                for _row in _ws.iter_rows(min_row=2, values_only=True):
                    if not _row or not _row[0] or str(_row[0]).strip() == "":
                        continue
                    _name   = str(_row[0]).strip()
                    _r      = max(0, min(255, int(float(_row[1] or 0))))
                    _g      = max(0, min(255, int(float(_row[2] or 0))))
                    _b      = max(0, min(255, int(float(_row[3] or 0))))
                    _area   = float(_row[4] or 0)
                    _custom = str(_row[5] or "").strip()
                    _preset = str(_row[6] or "[직접입력]").strip()
                    if _preset not in PRESET_OPTIONS:
                        _preset = "[직접입력]"
                    _new_rows.append({
                        "name": _name, "r": _r, "g": _g, "b": _b,
                        "preset": _preset, "custom_desc": _custom,
                        "area_sqm": _area, "tolerance": 25, "enabled": True,
                    })
                if _new_rows:
                    st.session_state.land_use_table = _new_rows
                    st.success("%d개 항목 불러옴" % len(_new_rows))
                    st.rerun()
                else:
                    st.warning("불러온 데이터가 없습니다.")
            except Exception as _e:
                st.error("엑셀 파일 오류: %s" % str(_e))

    # 색상 자동 추출 (DXF가 있으면 건너뜀)
    if (
        st.session_state.img_landuse_bytes
        and CV2_AVAILABLE
        and not st.session_state.dxf_layer_table
    ):

        # 이미 한 번 생성했는지 체크 (중복 방지)
        if not st.session_state.get("_auto_generated", False):

            with st.spinner("색상 및 면적 자동 계산 중..."):

                colors = extract_dominant_colors(
                    st.session_state.img_landuse_bytes,
                    n_colors=40
                )

                new_rows = []

                arr = np.array(bytes_to_pil(st.session_state.img_landuse_bytes))

                white_bg = (
                    (arr[:, :, 0] > 240) &
                    (arr[:, :, 1] > 240) &
                    (arr[:, :, 2] > 240)
                )

                # 흰색만 제외. 검정 도로 포함.
                valid_mask = ~white_bg
                total_px = max(1, int(np.count_nonzero(valid_mask)))
                for r, g, b, _ in colors:
                    r, g, b = int(r), int(g), int(b)
                    tol = 20
                    lo = np.array([max(0, r - tol), max(0, g - tol), max(0, b - tol)], dtype=np.uint8)
                    hi = np.array([min(255, r + tol), min(255, g + tol), min(255, b + tol)], dtype=np.uint8)
                    mask = cv2.inRange(arr, lo, hi)
                    mask = (mask > 0) & valid_mask
                    area_px = int(np.count_nonzero(mask))
                    ratio = area_px / total_px
                    if ratio < 0.003:
                        continue
                    area_sqm = float(st.session_state.site_area_sqm) * ratio
                    is_black = (r < 60 and g < 60 and b < 60)
                    new_rows.append({
                        "name": "도로" if is_black else "",
                        "r": r, "g": g, "b": b,
                        "preset": "[직접입력]" if is_black else "",
                        "custom_desc": "Road network, asphalt surface, lane markings, curb lines" if is_black else "",
                        "area_sqm": round(area_sqm, 1),
                        "tolerance": tol,
                        "enabled": True,
                    })

                if new_rows:
                    st.session_state.land_use_table = new_rows
                    st.session_state["_auto_generated"] = True
                    st.rerun()

    st.markdown("---")

    # 표준 항목 추가
    with st.expander("+ 항목 추가", expanded=False):
        lu_names = [lu[0] for lu in STANDARD_LAND_USES]
        ac1, ac2 = st.columns([3, 1])
        with ac1:
            sel_lu = st.selectbox("표준 항목에서 선택", lu_names)
        with ac2:
            if st.button("추가", type="primary"):
                idx = lu_names.index(sel_lu)
                item = STANDARD_LAND_USES[idx]
                st.session_state.land_use_table.append({
                    "name": item[0], "r": item[1][0], "g": item[1][1], "b": item[1][2],
                    "preset": item[2], "custom_desc": "",
                    "area_sqm": 10000.0, "tolerance": 25, "enabled": True,
                })
                st.rerun()

        st.markdown("**직접 입력**")
        dc1, dc2, dc3, dc4, dc5 = st.columns([2, 0.7, 0.7, 0.7, 1.5])
        new_name   = dc1.text_input("용도명", key="new_name")
        new_r      = dc2.number_input("R", 0, 255, 128, key="new_r")
        new_g      = dc3.number_input("G", 0, 255, 128, key="new_g")
        new_b      = dc4.number_input("B", 0, 255, 128, key="new_b")
        new_preset = dc5.selectbox("프리셋", PRESET_OPTIONS, key="new_preset")
        if st.button("직접 추가"):
            if new_name:
                st.session_state.land_use_table.append({
                    "name": new_name, "r": int(new_r), "g": int(new_g), "b": int(new_b),
                    "preset": new_preset, "custom_desc": "",
                    "area_sqm": 10000.0, "tolerance": 25, "enabled": True,
                })
                st.rerun()

    # 테이블 편집
    st.markdown('<div class="sub-label">토지이용 항목 목록</div>', unsafe_allow_html=True)
    table = st.session_state.land_use_table

    if not table:
        st.info("등록된 항목이 없습니다. 엑셀 업로드 또는 직접 추가하세요.")

    to_delete = []
    for i, row in enumerate(table):
        c_en, c_layer, c_name, c_r, c_g, c_b, c_tol, c_area, c_preset, c_chip, c_del = st.columns(
            [0.4, 1.6, 1.6, 0.6, 0.6, 0.6, 0.7, 1.1, 2.0, 0.5, 0.4]
        )
        r, g, b = int(row["r"]), int(row["g"]), int(row["b"])
        hex_color = "#%02x%02x%02x" % (r, g, b)

        table[i]["enabled"]   = c_en.checkbox("", value=row.get("enabled", True), key="en_%d" % i)
        table[i]["layer"]     = c_layer.text_input("", value=row.get("layer", ""), key="layer_%d" % i, label_visibility="collapsed", placeholder="DXF Layer")
        table[i]["name"]      = c_name.text_input("", value=row.get("name", ""), key="name_%d" % i, label_visibility="collapsed", placeholder="용도 입력")
        table[i]["r"]         = c_r.number_input("R", 0, 255, r, key="r_%d" % i, label_visibility="collapsed")
        table[i]["g"]         = c_g.number_input("G", 0, 255, g, key="g_%d" % i, label_visibility="collapsed")
        table[i]["b"]         = c_b.number_input("B", 0, 255, b, key="b_%d" % i, label_visibility="collapsed")
        table[i]["tolerance"] = c_tol.number_input("Tol", 5, 80, int(row.get("tolerance", 25)), key="tol_%d" % i, label_visibility="collapsed")
        table[i]["area_sqm"]  = c_area.number_input("sqm", 0.0, 9999999.0, float(row.get("area_sqm", 10000.0)), step=500.0, key="area_%d" % i, label_visibility="collapsed")

        cur_preset = row.get("preset", "")
        table[i]["preset"] = c_preset.selectbox(
            "", PRESET_OPTIONS, index=None if cur_preset == "" else PRESET_OPTIONS.index(cur_preset) if cur_preset in PRESET_OPTIONS else 0,
            key="preset_%d" % i, label_visibility="collapsed"
        )

        c_chip.markdown(
            '<div style="margin-top:6px;width:26px;height:26px;border-radius:5px;'
            'background:%s;border:1px solid rgba(0,0,0,0.2);"></div>' % hex_color,
            unsafe_allow_html=True
        )
        if c_del.button("x", key="del_%d" % i):
            to_delete.append(i)

        if table[i]["preset"] == "[직접입력]":
            table[i]["custom_desc"] = st.text_input(
                "용도 설명 (영문 권장)",
                value=row.get("custom_desc", ""),
                key="cdesc_%d" % i,
                placeholder="예: glamping resort with log cabins, 1~2F, natural wood facade",
                help="이 설명이 프롬프트에 직접 사용됩니다"
            )

    if to_delete:
        st.session_state.land_use_table = [r for idx, r in enumerate(table) if idx not in to_delete]
        st.rerun()

    # RGB 추출 미리보기
    st.markdown('<div class="sub-label">RGB 추출 미리보기</div>', unsafe_allow_html=True)
    if st.session_state.img_landuse_bytes and CV2_AVAILABLE and table:
        if st.button("구역 추출 미리보기", type="secondary"):
            with st.spinner("구역 추출 중..."):
                zm = extract_zone_masks(st.session_state.img_landuse_bytes, table)
            if zm:
                overlay = np.array(bytes_to_pil(st.session_state.img_landuse_bytes)).copy()
                total_px = sum(v["area_px"] for v in zm.values()) or 1
                for idx, z in zm.items():
                    cnts, _ = cv2.findContours(z["mask"], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cv2.drawContours(overlay, cnts, -1, (255, 0, 0), 3)
                    cx, cy = z["centroid"]
                    cv2.circle(overlay, (cx, cy), 8, (255, 0, 0), -1)
                    ratio = z["area_px"] / total_px * 100
                    name = table[idx].get("name", "Zone %d" % idx)
                    st.markdown("**%s**: %s px (%.1f%%) — %s"
                                % (name, "{:,}".format(z["area_px"]), ratio,
                                   describe_position(cx, cy, *z["img_size"])))
                st.image(Image.fromarray(overlay), caption="추출된 구역 (빨간 윤곽선)",
                         use_container_width=True)
            else:
                st.warning("추출된 구역 없음. RGB 값과 tolerance를 조정하세요.")

    # 면적 합계
    total_area = sum(row.get("area_sqm", 0) for row in table if row.get("enabled", True))
    site_area  = st.session_state.site_area_sqm
    diff = site_area - total_area
    diff_color = "#DC2626" if abs(diff) > site_area * 0.05 else "#16A34A"
    st.markdown(
        '<div style="background:#F0FDF4;border:1px solid #86EFAC;border-radius:8px;'
        'padding:12px 16px;margin-top:12px;">'
        '<b>면적 합계:</b> 개별 합계 <b>%s㎡</b> / 전체 부지 <b>%s㎡</b> / '
        '차이 <b style="color:%s;">%+.0f㎡</b></div>'
        % ("{:,.0f}".format(total_area), "{:,.0f}".format(site_area), diff_color, diff),
        unsafe_allow_html=True
    )

    st.success("설정 완료 시 '다음 ▶'로 이동하여 이미지를 생성하세요.")

# ══════════════════════════════════════════════════════════════
# STEP 2: 생성
# ══════════════════════════════════════════════════════════════
else:
    st.markdown('<div class="section-header">③ 조감도 생성</div>', unsafe_allow_html=True)

    if not GENAI_AVAILABLE:
        st.error("google-genai 패키지가 없습니다.")
        st.stop()

    if not st.session_state.dxf_landuse_hatch_bytes:
        st.error("landuse_hatch.png가 없습니다. STEP ①에서 DXF를 업로드하고 PNG 2장을 먼저 생성하세요.")
        st.stop()

    # DXF 기반 입력을 생성 파이프라인에 주입
    st.session_state.img_landuse_bytes = st.session_state.dxf_landuse_hatch_bytes
    if st.session_state.dxf_satellite_base_bytes:
        st.session_state.img_sat_bytes = st.session_state.dxf_satellite_base_bytes

    api_key = st.text_input("Google AI Studio API 키", type="password").strip()
    st.caption("모델: %s" % MODEL_NAME)

    table = st.session_state.land_use_table

    # 위성사진 상태
    if st.session_state.img_sat_bytes:
        st.success("✅ 위성사진 로드됨 — 합성 배경 및 경계 클립에 사용됩니다")
    else:
        st.warning("⚠️ 위성사진 없음 — 업로드하면 경계 클립 품질이 크게 향상됩니다")
        with st.expander("위성사진 업로드", expanded=True):
            f_sat_retry = st.file_uploader(
                "위성사진 (.png/.jpg/.jpeg)",
                type=["png","jpg","jpeg"],
                key="up_sat_retry"
            )
            if f_sat_retry:
                st.session_state.img_sat_bytes = f_sat_retry.getvalue()
                st.rerun()

    # 합성 입력 이미지 생성
    composite_bytes = None
    zone_label_map = {}
    site_mask = None

    if CV2_AVAILABLE:
        with st.spinner("입력 이미지 합성 중..."):
            composite_bytes, zone_label_map = build_composite_with_labels(
                st.session_state.img_landuse_bytes,
                table,
                st.session_state.img_sat_bytes
            )
            # 마스크 추출 (경계 클립용)
            site_mask = extract_site_mask(st.session_state.img_landuse_bytes)

    if composite_bytes:
        st.markdown('<div class="sub-label">합성 입력 이미지 (PASS1 입력)</div>',
                    unsafe_allow_html=True)
        st.image(bytes_to_pil(composite_bytes),
                 caption="위성배경 + 흰백판 + Z레이블 + 공원녹지 색상",
                 use_container_width=True)

    input_for_pass1 = composite_bytes if composite_bytes else st.session_state.img_landuse_bytes

    # 범례 이미지
    legend_bytes = build_legend_image(table)
    if legend_bytes:
        st.markdown('<div class="sub-label">범례 이미지 (UI 확인용)</div>', unsafe_allow_html=True)
        st.image(bytes_to_pil(legend_bytes), width=400,
                 caption="범례 — 모델 입력에는 사용되지 않습니다")

    # 프롬프트 생성
    zone_masks = {}
    if CV2_AVAILABLE:
        zone_masks = extract_zone_masks(st.session_state.img_landuse_bytes, table)

    pass1_prompt = build_pass1_prompt(
        table, zone_masks, st.session_state.site_area_sqm, zone_label_map
    )
    pass2_prompt = build_pass2_prompt(table)

    # 개발자 확인
    dev_pw = st.text_input("개발자 비밀번호", type="password", key="dev_pw")
    if dev_pw == "126791":
        with st.expander("PASS1 프롬프트", expanded=False):
            st.code(pass1_prompt, language="text")
        with st.expander("PASS2 프롬프트", expanded=False):
            st.code(pass2_prompt, language="text")
        with st.expander("Z 레이블 매핑", expanded=False):
            for z_key, row in zone_label_map.items():
                st.write("[%s] → %s (%s)" % (z_key, row.get("name", ""), row.get("preset", "")))

    st.markdown("---")

    # ── PASS 1 ────────────────────────────────────────────
    st.markdown("### STEP 1 — 2D 배치도 생성")

    # PASS1 이미지 입력: composite(Image1) + 위성(Image2)
    def get_pass1_images():
        imgs = [input_for_pass1]
        if st.session_state.img_sat_bytes:
            imgs.append(st.session_state.img_sat_bytes)
        return imgs

    def run_pass1():
        client = genai.Client(api_key=api_key)
        try:
            resp = client.models.generate_content(
                model=MODEL_NAME,
                contents=make_contents(pass1_prompt, get_pass1_images())
            )
            out = get_image_from_resp(resp)
            if out:
                # 후처리 1: 흰선 제거
                out = remove_white_lines(out)
                # 후처리 2: 경계 클립 (구조적 경계 이탈 차단)
                if site_mask is not None and st.session_state.img_sat_bytes:
                    out = apply_clip(out, st.session_state.img_sat_bytes, site_mask)
                st.session_state.pass1_outputs.append(out)
                st.session_state.pass1_selected_idx = len(st.session_state.pass1_outputs) - 1
                st.session_state.pass1_output_bytes = out
                st.session_state.pass2_outputs = []
                st.session_state.pass3_outputs = []
                st.session_state.pass2_output_bytes = None
            else:
                st.session_state._errors.append("PASS1: 이미지가 반환되지 않았습니다.")
        except Exception as e:
            st.session_state._errors.append("PASS1 오류: %s" % str(e))

    p1c1, p1c2, p1c3 = st.columns([1.2, 1.5, 3])
    with p1c1:
        if st.button("STEP 1 생성", type="primary", disabled=len(api_key) < 10):
            st.session_state.pass1_outputs = []
            st.session_state.pass1_selected_idx = 0
            with st.spinner("STEP 1 생성 중..."):
                run_pass1()
            st.rerun()
    with p1c2:
        if st.button("한 번 더 생성",
                     disabled=(len(api_key) < 10 or
                               not st.session_state.pass1_outputs or
                               len(st.session_state.pass1_outputs) >= 5)):
            with st.spinner("추가 생성 (%d번째)..." % (len(st.session_state.pass1_outputs) + 1)):
                run_pass1()
            st.rerun()
    with p1c3:
        if st.session_state.pass1_outputs:
            st.caption("%d개 생성됨 (최대 5개)" % len(st.session_state.pass1_outputs))

    if st.session_state.pass1_outputs:
        st.markdown("**STEP 1 결과 — 최적 이미지 선택**")
        selected_p1 = render_selector(st.session_state.pass1_outputs, "pass1_selected_idx", "STEP1")
        if selected_p1:
            st.session_state.pass1_output_bytes = selected_p1
        st.download_button("⬇️ STEP 1 다운로드",
                           data=st.session_state.pass1_output_bytes,
                           file_name="planvision_step1_2d.png", mime="image/png")

    st.markdown("---")

    # ── PASS 2 ────────────────────────────────────────────
    st.markdown("### STEP 2 — 3D 조감도 생성")
    p2_disabled = len(api_key) < 10 or not st.session_state.pass1_output_bytes

    def run_pass2():
        client = genai.Client(api_key=api_key)
        try:
            resp = client.models.generate_content(
                model=MODEL_NAME,
                contents=make_contents(pass2_prompt, [
                    st.session_state.pass1_output_bytes,
                    st.session_state.img_landuse_bytes,
                ])
            )
            out = get_image_from_resp(resp)
            if out:
                st.session_state.pass2_outputs.append(out)
                st.session_state.pass2_selected_idx = len(st.session_state.pass2_outputs) - 1
                st.session_state.pass2_output_bytes = out
                st.session_state.pass3_outputs = []
            else:
                st.session_state._errors.append("PASS2: 이미지가 반환되지 않았습니다.")
        except Exception as e:
            st.session_state._errors.append("PASS2 오류: %s" % str(e))

    p2c1, p2c2, p2c3 = st.columns([1.2, 1.5, 3])
    with p2c1:
        if st.button("STEP 2 생성", type="primary", disabled=p2_disabled):
            st.session_state.pass2_outputs = []
            st.session_state.pass2_selected_idx = 0
            with st.spinner("STEP 2 생성 중..."):
                run_pass2()
            st.rerun()
    with p2c2:
        if st.button("한 번 더 생성 ",
                     disabled=(p2_disabled or
                               not st.session_state.pass2_outputs or
                               len(st.session_state.pass2_outputs) >= 5)):
            with st.spinner("추가 생성 (%d번째)..." % (len(st.session_state.pass2_outputs) + 1)):
                run_pass2()
            st.rerun()
    with p2c3:
        if st.session_state.pass2_outputs:
            st.caption("%d개 생성됨 (최대 5개)" % len(st.session_state.pass2_outputs))

    if st.session_state.pass2_outputs:
        st.markdown("**STEP 2 결과 — 최적 이미지 선택**")
        selected_p2 = render_selector(st.session_state.pass2_outputs, "pass2_selected_idx", "STEP2")
        if selected_p2:
            st.session_state.pass2_output_bytes = selected_p2
        st.download_button("⬇️ STEP 2 다운로드",
                           data=st.session_state.pass2_output_bytes,
                           file_name="planvision_step2_3d.png", mime="image/png")

        st.markdown("---")

        # ── PASS 3 ────────────────────────────────────────
        st.markdown("### STEP 3 — 각도 변환 (선택사항)")
        p3_angle = st.selectbox("변환 각도",
                                ["30° (저각도)", "45° (중각도)", "60° (준조감도)"], index=1)
        angle_deg = {"30° (저각도)": 30, "45° (중각도)": 45, "60° (준조감도)": 60}[p3_angle]
        p3_disabled = len(api_key) < 10 or not st.session_state.pass2_output_bytes

        def run_pass3():
            client = genai.Client(api_key=api_key)
            try:
                resp = client.models.generate_content(
                    model=MODEL_NAME,
                    contents=make_contents(
                        build_pass3_prompt(angle_deg),
                        [st.session_state.pass2_output_bytes]
                    )
                )
                out = get_image_from_resp(resp)
                if out:
                    st.session_state.pass3_outputs.append(out)
                    st.session_state.pass3_selected_idx = len(st.session_state.pass3_outputs) - 1
                else:
                    st.session_state._errors.append("PASS3: 이미지가 반환되지 않았습니다.")
            except Exception as e:
                st.session_state._errors.append("PASS3 오류: %s" % str(e))

        p3c1, p3c2, p3c3 = st.columns([1.2, 1.5, 3])
        with p3c1:
            if st.button("STEP 3 변환", type="secondary", disabled=p3_disabled):
                st.session_state.pass3_outputs = []
                st.session_state.pass3_selected_idx = 0
                with st.spinner("STEP 3 생성 중..."):
                    run_pass3()
                st.rerun()
        with p3c2:
            if st.button("한 번 더 생성  ",
                         disabled=(p3_disabled or
                                   not st.session_state.pass3_outputs or
                                   len(st.session_state.pass3_outputs) >= 5)):
                with st.spinner("추가 생성 (%d번째)..." % (len(st.session_state.pass3_outputs) + 1)):
                    run_pass3()
                st.rerun()
        with p3c3:
            if st.session_state.pass3_outputs:
                st.caption("%d개 생성됨" % len(st.session_state.pass3_outputs))

        if st.session_state.pass3_outputs:
            st.markdown("**STEP 3 결과**")
            selected_p3 = render_selector(st.session_state.pass3_outputs, "pass3_selected_idx", "STEP3")
            p3_final = selected_p3 or st.session_state.pass3_outputs[0]
            st.download_button(
                "⬇️ 최종 다운로드 (STEP 3)", data=p3_final,
                file_name="planvision_final_%ddeg.png" % angle_deg,
                mime="image/png", use_container_width=True
            )
        else:
            st.download_button(
                "⬇️ 최종 다운로드 (STEP 2)",
                data=st.session_state.pass2_output_bytes,
                file_name="planvision_final_step2.png",
                mime="image/png"
            )
